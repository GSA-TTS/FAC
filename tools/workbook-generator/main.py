from collections import defaultdict
from collections import namedtuple as NT
from pathlib import Path
from playhouse.shortcuts import model_to_dict, dict_to_model
import os
import re
import sys

import argparse
import json
import pprint
import openpyxl as pyxl

pp = pprint.PrettyPrinter(indent=2)

# FIXME
# We'll want to figure out how to dynamically import the 
# models for a given year based on a command-line input, and
# then use that to match queries against the SQLite DBs.
# However, if we only use AY22 data, it won't matter.
from data.ay22.models import (
    Cfda,
    Gen,
    Passthrough,
    Findings,
    Findingstext,
    Ueis,
    Notes,
    Cpas,
    Captext
)

parser = argparse.ArgumentParser()

# This provides a way to map the sheet in the workbook to the 
# column in the DB. It also has a default value and 
# the type of value, so that things can be set correctly
# before filling in the XLSX workbooks.
FieldMap = NT('FieldMap', 'in_sheet in_db default type')

# tools/workbook-generator/templates/additional-ueis-workbook.xlsx 
# tools/workbook-generator/templates/audit-findings-text-workbook.xlsx 
# tools/workbook-generator/templates/corrective-action-plan-workbook.xlsx 
# tools/workbook-generator/templates/federal-awards-audit-findings-workbook.xlsx 
# tools/workbook-generator/templates/federal-awards-workbook.xlsx 
# tools/workbook-generator/templates/notes-to-sefa-workbook.xlsx 
# tools/workbook-generator/templates/PLACE-TEMPLATE-FILES-HERE 
# tools/workbook-generator/templates/secondary-auditors-workbook.xlsx

templates = {
    'AdditionalUEIs': 'additional-ueis-workbook.xlsx', 
    'AuditFindingsText': 'audit-findings-text-workbook.xlsx',
    'CAP': 'corrective-action-plan-workbook.xlsx',
    'AuditFindings': 'federal-awards-audit-findings-workbook.xlsx',
    'FederalAwards': 'federal-awards-workbook.xlsx',
    'SEFA': 'notes-to-sefa-workbook.xlsx',
    'SecondaryAuditors': 'secondary-auditors-workbook.xlsx'
}

def set_single_cell_range(wb, range_name, value):
    the_range = wb.defined_names[range_name]
    # The above returns a generator. Turn it to a list, and grab
    # the first element of the list. Now, this *tuple* contains a
    # sheet name and a cell reference... which you need to get rid
    # of the '$' to use.
    # https://itecnote.com/tecnote/python-using-excel-named-ranges-in-python-with-openpyxl/
    tup = list(the_range.destinations)[0]
    sheet_title = tup[0]
    cell_ref = tup[1].replace('$', '')
    ws = wb[sheet_title]
    ws[cell_ref] = value

# A tiny helper to index into workbooks.
# Assumes a capital letter.
def col_to_ndx(col):
    return ord(col) - 65 + 1

# Helper to set a range of values.
# Takes a named range, and then walks down the range,
# filling in values from the list past in (values).
def set_range(wb, range_name, values, default=None, type=str):
    the_range = wb.defined_names[range_name]
    dest = list(the_range.destinations)[0]
    sheet_title = dest[0]
    ws = wb[sheet_title]

    start_cell = dest[1].replace('$', '').split(':')[0]
    col = col_to_ndx(start_cell[0])
    start_row = int(start_cell[1])

    for ndx, v in enumerate(values):
        row = ndx+start_row
        if v:
            # This is a very noisy statement, showing everything
            # written into the workbook.
            # print(f'{range_name} c[{row}][{col}] <- {v} len({len(v)}) {default}')
            if v is not None:
                ws.cell(row=row, column=col, value=type(v))
            if len(v) == 0 and default is not None:
                # This is less noisy. Shows up for things like
                # empty findings counts. 2023 submissions
                # require that field to be 0, not empty,
                # if there are no findings.
                # print('Applying default')
                ws.cell(row=row, column=col, value=type(default))
        if not v:
            if default is not None:
               ws.cell(row=row, column=col, value=type(default))
            else:
                ws.cell(row=row, column=col, value='')
        else:
            # Leave it blank if we have no default passed in
            pass

def set_uei(wb, dbkey):
    g = Gen.select().where(Gen.dbkey == dbkey).get()
    set_single_cell_range(wb, 'auditee_uei', g.uei)
    return g

def map_simple_columns(wb, mappings, values):
    # Map all the simple ones
    for m in mappings:
        set_range(wb, 
                  m.in_sheet,
                  map(lambda v: model_to_dict(v)[m.in_db], values),
                  m.default, 
                  m.type)


# FIXME: Get the padding/shape right on the report_id
def dbkey_to_report_id(dbkey):
    g = Gen.select(Gen.audityear,Gen.fyenddate).where(Gen.dbkey == dbkey).get()
    month = g.fyenddate.split('-')[1]
    # 2022JUN0001000003
    # We start new audits at 1 million.
    # So, we want 10 digits, and zero-pad for 
    # historic DBKEY report_ids
    return f'{g.audityear}{month}{dbkey.zfill(10)}'

def generate_dissemination_test_table(api_endpoint, dbkey, mappings, objects):
    table = {
        'rows': list(),
        'singletons': dict()
    }
    table['endpoint'] = api_endpoint
    table['report_id'] = dbkey_to_report_id(dbkey)
    for o in objects:
        as_dict = model_to_dict(o)
        test_obj = {}
        test_obj['fields'] = []
        test_obj['values'] = []
        for m in mappings:
            # What if we only test non-null values?
            if ((m.in_db in as_dict) and as_dict[m.in_db] is not None) and (as_dict[m.in_db] != ""):               
                test_obj['fields'].append(m.in_sheet)
                test_obj['values'].append(as_dict[m.in_db])
        table['rows'].append(test_obj)
    return table


##########################################
#
# generate_findings
#
##########################################
def generate_findings(dbkey, outdir):
    print("--- generate findings ---")
    wb = pyxl.load_workbook(f'templates/{templates["AuditFindings"]}')
    mappings = [
       FieldMap('compliance_requirement', 'typerequirement', None, str), 
       FieldMap('reference_number', 'findingsrefnums', None, str),
       FieldMap('modified_opinion', 'modifiedopinion', None, str), 
       FieldMap('other_matters', 'othernoncompliance', None, str), 
       FieldMap('material_weakness', 'materialweakness', None, str), 
       FieldMap('significant_deficiency', 'significantdeficiency', None, str), 
       FieldMap('other_findings', 'otherfindings', None, str),
       FieldMap('questioned_costs', 'qcosts', None, str),
       FieldMap('repeat_prior_reference', 'repeatfinding', None, str), 
       FieldMap('prior_references', 'priorfindingrefnums', None, str), 
       # is_valid is computed in the workbook
    ]
    g = set_uei(wb, dbkey)
    cfdas = Cfda.select(Cfda.elecauditsid).where(Cfda.dbkey == g.dbkey)
    findings = Findings.select().where(Findings.dbkey == g.dbkey)

    map_simple_columns(wb, mappings, findings)

    # For each of them, I need to generate an elec -> award mapping.
    e2a = {}
    award_references = []
    if (cfdas != None) and (findings != None):
        for ndx, cfda in enumerate(cfdas):
            if cfda:
                e2a[cfda.elecauditsid] = f'AWARD-{ndx+1:04d}'

        #award_references = []
        if len(e2a) != 0:
            for find in findings:
                if find:
                    award_references.append(e2a[find.elecauditsid])

            if len(award_references) != 0:
                # print("award_references",  award_references)
                set_range(wb, 'award_reference', award_references)
    
    wb.save(os.path.join(outdir, f'findings-{dbkey}.xlsx'))
    
    table = generate_dissemination_test_table('findings', dbkey, mappings, findings)
    # Add the award references to the objects.
    for obj, ar in zip(table['rows'], award_references):
        obj['fields'].append('award_reference')
        obj['values'].append(ar)

    return table

##########################################
#
# generate_federal_awards
#
##########################################
def generate_federal_awards(dbkey, outdir):
    print("--- generate federal awards ---")
    wb = pyxl.load_workbook(f'templates/{templates["FederalAwards"]}')
    # In sheet : in DB
    mappings = [
        FieldMap('program_name', 'federalprogramname', None, str),
        FieldMap('additional_award_identification', 'awardidentification', None, str),
        FieldMap('cluster_name', 'clustername', "N/A", str),
        FieldMap('state_cluster_name', 'stateclustername', None, str),
        FieldMap('other_cluster_name', 'otherclustername', None, str),
        FieldMap('federal_program_total', 'programtotal', 0, int),
        FieldMap('cluster_total', 'clustertotal', 0, float),
        FieldMap('is_guaranteed', 'loans', None, str),
        FieldMap('loan_balance_at_audit_period_end', 'loanbalance', None, float),
        FieldMap('is_direct', 'direct', None, str),
        FieldMap('is_passed', 'passthroughaward', None, str),
        FieldMap('subrecipient_amount', 'passthroughamount', None, float),
        FieldMap('is_major', 'majorprogram', None, str),
        FieldMap('audit_report_type', 'typereport_mp', None, str),
        FieldMap('number_of_audit_findings', 'findings', 0, int),
        FieldMap('amount_expended', 'amount', 0, int),
        FieldMap('federal_program_total', 'programtotal', 0, int)
    ]
    g = set_uei(wb, dbkey)
    cfdas = Cfda.select().where(Cfda.dbkey == g.dbkey)
    map_simple_columns(wb, mappings, cfdas)
    
    # Map things with transformations
    prefixes = map(lambda v: (v.cfda).split('.')[0], cfdas)
    extensions = map(lambda v: (v.cfda).split('.')[1], cfdas)
    set_range(wb, 'federal_agency_prefix', prefixes)
    set_range(wb, 'three_digit_extension', extensions)

    # We have to hop through several tables to build a list 
    # of passthrough names. Note that anything without a passthrough
    # needs to be represented in the list as an empty string.
    passthrough_names = []
    passthrough_ids = []
    for cfda in cfdas:
        try:
            pnq = (Passthrough
                   .select()
                   .where((Passthrough.dbkey == cfda.dbkey) &
                          (Passthrough.elecauditsid == cfda.elecauditsid))).get()
            passthrough_names.append(pnq.passthroughname)
            passthrough_ids.append(pnq.passthroughid)
            # print(f'Looking up {cfda.dbkey}, {cfda.elecauditsid} <- {json.dumps(model_to_dict(pnq))}')
        except Exception as e:
            passthrough_names.append('')
            passthrough_ids.append('')
    set_range(wb, 'passthrough_name', passthrough_names)
    set_range(wb, 'passthrough_identifying_number', passthrough_ids)

    # Total amount expended must be calculated and inserted
    total = 0
    for cfda in cfdas:
        total += int(cfda.amount)
    set_single_cell_range(wb, 'total_amount_expended', total)

    wb.save(os.path.join(outdir, f'federal-awards-{dbkey}.xlsx'))

    table = generate_dissemination_test_table('federal_awards', dbkey, mappings, cfdas)
    # prefix
    for obj, pfix, ext in zip(table['rows'], prefixes, extensions):
        obj['fields'].append('federal_agency_prefix')
        obj['values'].append(pfix)
        obj['fields'].append('three_digit_extension')
        obj['values'].append(ext)
    # names, ids
    for obj, name, id in zip(table['rows'], passthrough_names, passthrough_ids):
        obj['fields'].append('passthrough_name')
        obj['values'].append(name)
        obj['fields'].append('passthrough_identifying_number')
        obj['values'].append(id)
    table['singletons']['auditee_uei'] = g.uei
    table['singletons']['total_amount_expended'] = total

    return table

##########################################
#
# generate_findings_text
#
##########################################
def generate_findings_text(dbkey, outdir):
    print("--- generate findings text ---")
    wb = pyxl.load_workbook(f'templates/{templates["AuditFindingsText"]}')
    mappings = [
        FieldMap('reference_number', 'findingrefnums', None, str),
        FieldMap('text_of_finding', 'text', None, str),
        FieldMap('contains_chart_or_table', 'chartstables', None, str),
    ]
    g = set_uei(wb, dbkey)
    ftexts = Findingstext.select().where(Findingstext.dbkey == g.dbkey)
    map_simple_columns(wb, mappings, ftexts)
    wb.save(os.path.join(outdir, f'findings-text-{dbkey}.xlsx'))
    table = generate_dissemination_test_table('findings_text', dbkey, mappings, ftexts)
    table['singletons']['auditee_uei'] = g.uei

    return table

##########################################
#
# generate_additional_ueis
#
##########################################
def generate_additional_ueis(dbkey, outdir):
    print("--- generate additional ueis ---")
    wb = pyxl.load_workbook(f'templates/{templates["AdditionalUEIs"]}')
    mappings = [
        FieldMap('additional_uei', 'uei', None, str),
        #FieldMap('ueiseqnum', 'ueiseqnum', 0, int)
    ]

    g = set_uei(wb, dbkey)
    addl_ueis = Ueis.select().where(Ueis.dbkey == g.dbkey)
    map_simple_columns(wb, mappings, addl_ueis)
    wb.save(os.path.join(outdir, f'additional-ueis-{dbkey}.xlsx'))
    table = generate_dissemination_test_table('additional_ueis', dbkey, mappings, addl_ueis)
    table['singletons']['auditee_uei'] = g.uei

    return table

##########################################
#
# generate_notes_to_sefa
#
##########################################
def generate_notes_to_sefa(dbkey, outdir):
    print("--- generate notes to sefa ---")
    wb = pyxl.load_workbook(f'templates/{templates["SEFA"]}')
    mappings = [
        FieldMap('note_title', 'title', None, str),
        FieldMap('note_content', 'content', None, str)
    ]
    g =  set_uei(wb, dbkey)
    # The mapping is weird.
    # https://facdissem.census.gov/Documents/DataDownloadKey.xlsx
    # The TYPEID column determines which field in the form a given row corresponds to.
    # TYPEID=1 is the description of significant accounting policies.
    # TYPEID=2 is the De Minimis cost rate.
    # TYPEID=3 is for notes, which have sequence numbers... that must align somewhere.
    policies = Notes.select().where((Notes.dbkey == g.dbkey) & (Notes.type_id == 1)).get()
    rate = Notes.select().where((Notes.dbkey == g.dbkey) & (Notes.type_id == 2)).get()
    notes = Notes.select().where((Notes.dbkey == g.dbkey) & (Notes.type_id == 3)).order_by(Notes.seq_number)

    # This looks like the right way to set the three required fields
    set_single_cell_range(wb, 'accounting_policies', policies.content)
    # WARNING
    # This is being faked. We're askign a Y/N question in the collection. 
    # Census just let them type some stuff. So, this is a rough
    # attempt to generate a Y/N value from the content. 
    # This means the data is *not* true to what was intended, but
    # it *is* good enough for us to use for testing.
    is_used = "Huh"
    if (re.search('did not use', rate.content)
        or re.search('not to use', rate.content)
        or re.search('not use', rate.content)
        or re.search('not elected', rate.content)
    ):
        is_used = "N"
    elif re.search("used", rate.content):
        is_used = "Y"
    else:
        is_used = "Y&N"

    set_single_cell_range(wb, 'is_minimis_rate_used', is_used)
    set_single_cell_range(wb, 'rate_explained', rate.content)
    
    # Map the rest as notes.
    map_simple_columns(wb, mappings, notes)
    wb.save(os.path.join(outdir, f'notes-{dbkey}.xlsx'))

    table = generate_dissemination_test_table('notes_to_sefa', dbkey, mappings, notes)
    table['singletons']['accounting_policites'] = policies.content
    table['singletons']['is_minimis_rate_used'] =  is_used
    table['singletons']['rate_explained'] = rate.content
    table['singletons']['auditee_uei'] = g.uei
    return table

##########################################
#
# generate_secondary_auditors
#
##########################################
def generate_secondary_auditors(dbkey, outdir):
    print("--- generate secondary auditors ---")
    wb = pyxl.load_workbook(f'templates/{templates["SecondaryAuditors"]}')
    mappings = [
        FieldMap('secondary_auditor_address_city', 'cpacity', None, str),
        FieldMap('secondary_auditor_contact_name', 'cpacontact', None, str),
        FieldMap('secondary_auditor_ein', 'cpaein', 0, int),
        FieldMap('secondary_auditor_contact_email', 'cpaemail', None, str),
        FieldMap('secondary_auditor_name', 'cpafirmname', None, str),
        FieldMap('secondary_auditor_contact_phone', 'cpaphone', None, str),
        FieldMap('secondary_auditor_address_state', 'cpastate', None, str),
        FieldMap('secondary_auditor_address_street', 'cpastreet1', None, str),
        FieldMap('secondary_auditor_contact_title', 'cpatitle', None, str),
        FieldMap('secondary_auditor_address_zipcode', 'cpazipcode', None, str)
    ]

    g = set_uei(wb, dbkey)
    sec_cpas = Cpas.select().where(Cpas.dbkey == g.dbkey)
    
    map_simple_columns(wb, mappings, sec_cpas)
    wb.save(os.path.join(outdir, f'cpas-{dbkey}.xlsx'))
    
    table = generate_dissemination_test_table('secondary_auditors', dbkey, mappings, sec_cpas)
    table['singletons']['auditee_uei'] = g.uei

    return table

##########################################
#
# generate_captext
#
##########################################
def generate_captext(dbkey, outdir):
    print("--- generate corrective action plan ---")
    wb = pyxl.load_workbook(f'templates/{templates["CAP"]}')
    mappings = [
        FieldMap('reference_number', 'findingrefnums', None, str),
        FieldMap('planned_action', 'text', None, str),
        FieldMap('contains_chart_or_table', 'chartstables', None, str)
    ]
    
    g = set_uei(wb, dbkey)
    captexts = Captext.select().where(Captext.dbkey == g.dbkey)
    
    map_simple_columns(wb, mappings, captexts)
    wb.save(os.path.join(outdir, f'captext-{dbkey}.xlsx'))

    table = generate_dissemination_test_table('cap_text', dbkey, mappings, captexts)
    table['singletons']['auditee_uei'] = g.uei

    return table


##########################################
def main():
    out_basedir = None
    if args.output:
        out_basedir = args.output
    else:
        out_basedir = 'output'

    if not os.path.exists(out_basedir):
        try:
            os.mkdir(out_basedir)
        except Exception as e:
            pass
    outdir = os.path.join(out_basedir, args.dbkey)
    if not os.path.exists(outdir):
        try:
            os.mkdir(outdir)
        except Exception as e:
            print('could not create output directory. exiting.')
            sys.exit()

    fat = generate_federal_awards(args.dbkey, outdir)
    ft = generate_findings(args.dbkey, outdir) 
    ftt = generate_findings_text(args.dbkey, outdir)
    aut = generate_additional_ueis(args.dbkey, outdir)
    ntst = generate_notes_to_sefa(args.dbkey, outdir)
    sat = generate_secondary_auditors(args.dbkey, outdir)
    ctt = generate_captext(args.dbkey, outdir)
    tables = [sat, ctt, ntst, aut, ftt, ft, fat]
    with open(os.path.join(outdir, f'test-array-{args.dbkey}.json'), "w") as test_file:
        jstr = json.dumps(tables, indent=2, sort_keys=True)
        test_file.write(jstr)
    
if __name__ == '__main__':
    parser.add_argument('--dbkey', type=str, required=True)
    parser.add_argument('--output', type=str, required=False)
    args = parser.parse_args()
    main()
