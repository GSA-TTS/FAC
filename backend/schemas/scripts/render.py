import os
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname, absolute_coordinate
from openpyxl.styles import PatternFill, Alignment, Protection, Font

from collections import namedtuple as NT

# from openpyxl.styles.colors import Color
# from openpyxl.worksheet.dimensions import ColumnDimension

import json

# from json import JSONEncoder
import _jsonnet

import sys

from xkcdpass import xkcd_password as xp
from openpyxl.workbook.protection import WorkbookProtection

import parse

Range = NT(
    "Range",
    "name,column,label_row,range_start_row,range_start,abs_range_start,full_range",
)

# We can compile larger workbooks by setting
# an environment variable for one-offs.
DEFAULT_MAX_ROWS = 10000
rows = os.getenv("MAX_ROWS", DEFAULT_MAX_ROWS)
try:
    MAX_ROWS = int(rows)
except ValueError:
    MAX_ROWS = DEFAULT_MAX_ROWS

XLSX_MAX_ROWS = 1048576  # Excel has a maximum of 1048576 rows
XLSX_MAX_COLS = 16384  # Excel has a maximum of 16384 columns

# Styling
header_row_fill = PatternFill(
    fill_type="solid",
    start_color="000066CC",
    end_color="000066CC",
    bgColor="000066CC",
    patternType="solid",
)

meta_row_fill = PatternFill(
    fill_type="solid",
    start_color="333333",
    end_color="333333",
    bgColor="333333",
    patternType="solid",
)

header_row_font = Font(
    name="Calibri",
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline="none",
    strike=False,
    color="00FFFFFF",
)

meta_row_font = Font(
    name="Calibri",
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline="none",
    strike=False,
    color="00FFFFFF",
)


def process_spec(WBNT):
    """
    process_spec is the start of the process. It creates an
    empty workbook and proceeds to build up the final result through
    a series of steps that apply changes to the workbook object.
    """
    wb = create_empty_workbook()
    password = generate_password()
    for ndx, sheet in enumerate(WBNT.sheets):
        print("########################")
        print(f"## Processing sheet {ndx+1}")
        print("########################")
        ws = create_protected_sheet(wb, sheet, password, ndx)
        if sheet.hide_col_from is not None:
            hide_all_columns_from(ws, sheet.hide_col_from)
        if sheet.hide_row_from is not None:
            hide_all_rows_from(ws, sheet.hide_row_from)
        process_open_ranges(wb, ws, sheet)
        add_validations(wb, ws, sheet.open_ranges)
        add_validations(wb, ws, sheet.text_ranges)
        activate_wraptext(wb)
        apply_formula(ws, WBNT.title_row + 1, sheet)
        process_single_cells(wb, ws, sheet)
        process_meta_cells(wb, ws, sheet)
        process_text_ranges(wb, ws, sheet)
        unlock_data_entry_cells(WBNT.title_row, ws, sheet)
        set_column_widths(wb, ws, sheet)
        set_row_heights(wb, ws, sheet)

    set_wb_security(wb, password)
    return wb


def create_empty_workbook():
    """
    Creates an empty workbook for population.
    """
    wb = Workbook()
    ws = wb.active
    # Remove the default/active worksheet. We'll make more.
    wb.remove(ws)
    return wb


def generate_password():
    """
    Generates a long password that we use to lock the workbook.
    We currently assume we will never unlock the workbook, so we ultimately
    throw this away/ignore it.
    """
    wordfile = xp.locate_wordfile()
    words = xp.generate_wordlist(wordfile=wordfile, min_length=7, max_length=9)
    correct_horse = "-".join(xp.generate_xkcdpassword(words).split(" "))
    return correct_horse


def create_protected_sheet(wb, sheet, password, ndx):
    """
    We typically have one sheet per workbook, but some have enumerated values
    that we want to check against. Each sheet must have password protection set
    independently.
    """
    print("---- created_protected_sheet ----")
    ws = wb.create_sheet(title=sheet.name, index=ndx)
    # The sheet has to be locked, and individual cells unlocked.
    # https://stackoverflow.com/questions/46877091/lock-some-cells-from-editing-in-python-openpyxl
    ws.protection.set_password(password)
    ws.protection.sheet = True
    ws.protection.enable()
    return ws


def hide_all_columns_from(ws, start_column):
    """Hides all columns from the specified column to the end of the sheet."""
    for col_idx in range(start_column, XLSX_MAX_COLS + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].hidden = True


def hide_all_rows_from(ws, start_row):
    """Hides all rows from the specified row to the end of the sheet."""
    for row_idx in range(start_row, XLSX_MAX_ROWS + 1):
        ws.row_dimensions[row_idx].hidden = True


def apply_cell_format(cell, format):
    """
    Applies formatting to the cell based on the format defined in the spec.
    """
    if format == "text":
        cell.number_format = "@"
    elif format == "dollar":
        cell.number_format = "$#,##0"


def apply_range_format(ws, coords, format):
    """
    Applies formatting to the range of cells based on the format defined in the spec.
    """
    for rowndx in range(coords.range_start_row, MAX_ROWS + 1):
        cell_reference = f"{coords.column}{rowndx}"
        cell = ws[cell_reference]
        apply_cell_format(cell, format)


def process_open_ranges(wb, ws, sheet):
    """
    Open ranges are the column-wise data in the sheet. They are called "open" ranges
    because they go from a given start point "all the way down," and therefore are
    open-ended.
    """
    print(f"---- process_open_ranges `{sheet.name}` ----")
    for r in sheet.open_ranges:
        coords = make_range(r)
        cell_reference = f"{quote_sheetname(ws.title)}!{coords.full_range}"
        print(
            f"Open Range: {r.posn.title}, {coords.column}, {coords.abs_range_start}, {cell_reference}"
        )
        new_range = DefinedName(name=coords.name, attr_text=cell_reference)
        wb.defined_names.add(new_range)
        configure_header_cell(ws, r)
        if r.posn.format is not None:
            apply_range_format(ws, coords, r.posn.format)
        # Make the header row tall.
        ws.row_dimensions[coords.range_start_row - 1].height = (
            sheet.header_height if sheet.header_height else 100
        )


#########################################################
# VALIDATIONS


def add_validations(wb, ws, ranges):
    """
    Embedded validations do a lot of work, but also have a lot of detail
    in how they are processed and attached. Configure does most of the work.
    """
    for r in ranges:
        coords = make_range(r)
        configure_validation(wb, ws, coords, r)


def configure_validation(wb, ws, coords: Range, r):
    print("---- configure_validations ----")
    dv = None

    if r.validation.type == "NOVALIDATION":
        print("\t ---- NOVALIDATION")
        return  # Exit early, no further action needed

    print("----", r.validation.type)

    if r.validation.type == "list":
        dv = DataValidation(
            type="list",
            formula1=r.validation.formula1.format(coords.range_start_row),
            allow_blank=True,
            errorStyle=r.validation.errorStyle,
        )
        # https://stackoverflow.com/questions/75889368/openpyxl-excel-file-created-is-not-showing-validation-errors-or-prompt-message
    elif r.validation.type == "range_lookup":
        # https://openpyxl.readthedocs.io/en/latest/validation.html#:~:text=Cell%20range%20validation
        # rng = wb.defined_names[f"{r.validation.lookup_range}"]
        dv = DataValidation(
            type="list", formula1=r.validation.lookup_range, allow_blank=True
        )
        print(f"---- formula1: {r.validation.lookup_range}")
    elif r.validation.type in ["custom", "lookup"]:
        dv = DataValidation(type="custom", allow_blank=True)
        dv.formula1 = r.validation.formula1
        # This to handle the cases where formula1 is expressed with placeholders {0}.
        dv.formula1 = dv.formula1.format(coords.range_start_row)
        if "FIRSTCELLREF" in dv.formula1:
            dv.formula1 = dv.formula1.replace(
                "FIRSTCELLREF", f"${coords.column}{coords.range_start_row}"
            )
        if "LASTCELLREF" in dv.formula1:
            dv.formula1 = dv.formula1.replace(
                "LASTCELLREF", f"${coords.column}${MAX_ROWS}"
            )
        if "LOOKUPRANGE" in dv.formula1:
            dv.formula1 = dv.formula1.replace("LOOKUPRANGE", r.validation.lookup_range)
    elif r.validation.type == "textLength":
        dv = DataValidation(
            type="textLength",
            formula1=r.validation.formula1,
            operator=r.validation.operator,
            allow_blank=True,
        )

    # Properties attached to the validation object
    if dv:
        dv.showErrorMessage = True
        dv.operator = r.validation.operator or dv.operator
        dv.allow_blank = r.validation.allow_blank or dv.allow_blank
        dv.error = r.validation.custom_error or dv.error
        dv.errorTitle = r.validation.custom_title or dv.errorTitle

        dv.showDropDown = False
        ws.add_data_validation(dv)
        dv.showErrorMessage = True
        dv.add(coords.full_range)


def add_yorn_validation(ws):
    yorn_dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    yorn_dv.error = "Entries must be Y or N in this column"
    yorn_dv.errorTitle = "Must by Y or N"
    yorn_dv.showDropDown = False
    # https://stackoverflow.com/questions/75889368/openpyxl-excel-file-created-is-not-showing-validation-errors-or-prompt-message
    yorn_dv.showErrorMessage = True
    ws.add_data_validation(yorn_dv)
    return yorn_dv


def apply_formula(ws, data_row, sheet):
    print("---- apply_formula ----")
    # Apply formulas to the open ranges
    # FIXME MCJ: I don't know if semantics were preserved in my update.
    for r in sheet.open_ranges:
        coords = make_range(r)
        if r.formula is not None:
            for row in range(coords.range_start_row, MAX_ROWS + 1):
                ws[f"{coords.column}{row}"] = str((r.formula)).format(row)

    # Apply formulas to the single cells
    for r in sheet.single_cells:
        if r.formula is not None:
            formula = r.formula
            if "FIRSTCELLREF" in formula:
                formula = formula.replace(
                    "FIRSTCELLREF", f"{r.posn.range_cell[0]}{data_row}"
                )
            if "LASTCELLREF" in formula:
                formula = formula.replace(
                    "LASTCELLREF", f"{r.posn.range_cell[0]}{MAX_ROWS}"
                )
            if "FIRSTROW" in formula:
                formula = formula.replace("FIRSTROW", f"{data_row}")
            if "LASTROW" in formula:
                formula = formula.replace("LASTROW", f"{MAX_ROWS}")
            print("FORMULA")
            print(f"{r.posn.range_cell} :: {formula}")
            ws[r.posn.range_cell] = formula


######################################################
# SHEET LOADING


def jsonnet_sheet_spec_to_json(filename):
    json_str = _jsonnet.evaluate_snippet(filename, open(filename).read())
    jobj = json.loads(json_str)
    return jobj


######################################################
# SHEET BUILDING


def process_single_cells(wb, ws, sheet):
    print("---- process_single_cells ----")
    # Create all the single cells
    for o in sheet.single_cells:
        cell_coordinate = o.posn.range_cell
        absolute_cell_coordinate = f"{absolute_coordinate(cell_coordinate)}"
        cell_row = int(o.posn.range_cell[1])
        cell_column = o.posn.range_cell[0]
        sheet_cell_coordinate = (
            f"{quote_sheetname(ws.title)}!{absolute_cell_coordinate}"
        )
        print(f"Single Cell: {absolute_cell_coordinate} {sheet_cell_coordinate}")
        new_range = DefinedName(name=o.posn.range_name, attr_text=sheet_cell_coordinate)
        wb.defined_names.add(new_range)
        the_cell = ws[o.posn.title_cell]
        the_cell.value = o.posn.title
        the_cell.fill = header_row_fill
        the_cell.font = header_row_font
        the_cell.alignment = Alignment(wrapText=True, wrap_text=True)
        # ws.row_dimensions[cell_row].height = 40
        entry_cell_obj = ws[absolute_cell_coordinate]
        if not o.posn.keep_locked:
            entry_cell_obj.protection = Protection(locked=False)
        # Creating a coord for the single-cell range for validation configuration
        coord = Range(
            "",
            cell_column,
            0,
            cell_row,
            absolute_cell_coordinate,
            "",
            f"{o.posn.range_cell}:{o.posn.range_cell}",
        )
        configure_validation(wb, ws, coord, o)
        if o.posn.format is not None:
            apply_cell_format(entry_cell_obj, o.posn.format)

        if sheet.header_height:
            row = int(o.posn.title_cell[1])
            ws.row_dimensions[row].height = sheet.header_height

        if o.value:
            cell_reference = o.posn.range_cell
            ws[cell_reference] = o.value


def process_meta_cells(wb, ws, sheet):
    print("---- process_meta_cells ----")
    # Create all the meta cells
    for o in sheet.meta_cells:
        cell_coordinate = o.posn.title_cell
        absolute_cell_coordinate = f"{absolute_coordinate(cell_coordinate)}"
        sheet_cell_coordinate = (
            f"{quote_sheetname(ws.title)}!{absolute_cell_coordinate}"
        )
        print(f"Meta Cell: {absolute_cell_coordinate} {sheet_cell_coordinate}")
        the_cell = ws[o.posn.title_cell]
        the_cell.value = o.posn.title
        the_cell.fill = meta_row_fill
        the_cell.font = meta_row_font
        the_cell.alignment = Alignment(wrapText=True, wrap_text=True)
        entry_cell_obj = ws[absolute_cell_coordinate]
        if not o.posn.keep_locked:
            entry_cell_obj.protection = Protection(locked=False)
        if o.posn.format is not None:
            apply_cell_format(entry_cell_obj, o.posn.format)

        if sheet.header_height:
            row = int(o.posn.title_cell[1])
            ws.row_dimensions[row].height = sheet.header_height


def make_range(r):
    column = r.posn.title_cell[0]
    title_row = int(r.posn.title_cell[1])
    range_start_row = int(r.posn.title_cell[1]) + 1
    start_cell = column + str(range_start_row)
    last_row = (
        r.posn.last_range_cell[1:] if r.posn.last_range_cell is not None else MAX_ROWS
    )
    full_range = f"${column}${range_start_row}:${column}${last_row}"
    return Range(
        r.posn.range_name,
        column,
        title_row,
        range_start_row,
        start_cell,
        f"{absolute_coordinate(start_cell)}",
        full_range,
    )


def configure_header_cell(ws, r):
    the_cell = ws[r.posn.title_cell]
    the_cell.value = r.posn.title
    the_cell.fill = header_row_fill
    the_cell.font = header_row_font
    the_cell.alignment = Alignment(wrapText=True, wrap_text=True)


def process_text_ranges(wb, ws, sheet):
    print("---- parse_text_ranges ----")
    max_width = 72
    for tr in sheet.text_ranges:
        coords = make_range(tr)
        cell_reference = f"{quote_sheetname(ws.title)}!{coords.full_range}"
        ws[tr.posn.title_cell] = tr.posn.title
        start_cell_column = tr.posn.title_cell[0]
        start_cell_row = int(tr.posn.title_cell[1]) + 1
        for ndx, v in enumerate(tr.contents.values):
            ws[f"{start_cell_column}{start_cell_row + ndx}"] = v
            if len(v) > max_width:
                max_width = len(v)
        new_range = DefinedName(name=coords.name, attr_text=cell_reference)
        print(f"Adding named text range: {coords.name}, {cell_reference}")
        wb.defined_names.add(new_range)

        ws.column_dimensions[start_cell_column].width = max_width


def unlock_data_entry_cells(header_row, ws, sheet):
    print("---- unlock_data_entry_cells ----")
    for r in sheet.open_ranges:
        if not r.posn.keep_locked:
            coords = make_range(r)
            for rowndx in range(coords.range_start_row, MAX_ROWS + 1):
                cell_reference = f"${coords.column}${rowndx}"
                cell = ws[cell_reference]
                cell.protection = Protection(locked=False)


def calculate_average_width(ws):
    if len(list(ws.rows)) == 0:
        return 72, {}
    else:
        widths = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    widths[cell.column_letter] = max(
                        (widths.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        sum_widths = sum(widths.values())
        avg_width = sum_widths / len(widths)
        return avg_width, widths


def set_column_dimensions(ws, sheet, avg_width, widths):
    for col, value in widths.items():
        if col not in [o.posn.title_cell[0] for o in sheet.single_cells]:
            ws.column_dimensions[col].width = avg_width / 2


def set_open_range_and_single_cell_widths(ws, sheet):
    for r in sheet.open_ranges:
        column = r.posn.title_cell[0]
        if r.posn.width:
            ws.column_dimensions[column].width = r.posn.width
    for r in sheet.single_cells:
        column = r.posn.title_cell[0]
        if r.posn.width:
            ws.column_dimensions[column].width = r.posn.width


def set_row_heights(wb, ws, sheet):
    """Set the row heights based on sheet.row_height if defined."""
    if sheet.row_height and ws.title == "Form":
        named_cells = set()  # Set to store all cells within named ranges

        # Collecting all cells that are in named ranges
        for named_range in wb.defined_names:
            for dest in wb.defined_names[
                named_range
            ].destinations:  # Each destination is a tuple of (sheet_name, cell_range)
                if dest[0] == "Form":
                    for row in ws[dest[1]]:
                        for cell in row:
                            named_cells.add((cell.row, cell.column))

        # Checking and setting the height for each cell in named ranges
        for cell_row, cell_col in named_cells:
            ws.row_dimensions[cell_row].height = sheet.row_height


def activate_wraptext(wb):
    for ws in wb.worksheets:
        if ws.title == "Form":
            for named_range_name in wb.defined_names:
                activate_wraptext_for_named_range(wb, ws, named_range_name)


def activate_wraptext_for_named_range(wb, ws, named_range_name):
    """Activate wrapText for all cells within a named range in the worksheet."""
    if named_range_name in wb.defined_names:
        named_range = wb.defined_names[named_range_name]
        if named_range.attr_text.startswith("'Form'!"):
            for cell_range in named_range.destinations:
                for row in ws[cell_range[1]]:
                    for cell in row:
                        cell.alignment = Alignment(wrapText=True)


def set_column_widths(wb, ws, sheet):
    # Set he widths to something... sensible.
    # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    print("---- set_column_widths ----")
    avg_width, widths = calculate_average_width(ws)
    set_column_dimensions(ws, sheet, avg_width, widths)
    set_open_range_and_single_cell_widths(ws, sheet)


def set_wb_security(wb, password):
    # This is not doing what I want.
    # I cannot prevent the sheets from unlocking.
    wb.security = WorkbookProtection(workbookPassword=password, lockStructure=True)
    wb.security.lockStructure = True
    print(f"To unlock: {password}")


def save_workbook(wb, basename):
    wb.save(f"{basename}")


if __name__ == "__main__":
    print("render.py")
    if len(sys.argv) == 3:
        maybe_filename = sys.argv[1]
        if maybe_filename.endswith("jsonnet"):
            spec = jsonnet_sheet_spec_to_json(maybe_filename)
            parsed = parse.parse_spec(spec)
            # `parsed` is going to be a WB NT.
            wb = process_spec(parsed)
            # os.path.splitext(os.path.basename(sys.argv[2]))[0])
            save_workbook(wb, sys.argv[2])
