import json
import shutil
from pathlib import Path

from openpyxl import load_workbook


SCHEMAS_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = SCHEMAS_DIR / "output" / "excel" / "xlsx"
OUTPUT_DIR = SCHEMAS_DIR.parent / "cypress" / "fixtures" / "test_workbooks"
FIXTURES_FILE = SCHEMAS_DIR / "source" / "data" / "test_workbook_fixtures.json"


def parse_cell_ref(cell_ref: str) -> tuple[str, str]:
    if "!" not in cell_ref:
        raise ValueError(f"Invalid cell reference '{cell_ref}'. Expected 'Sheet!Cell'.")
    sheet_name, cell = cell_ref.split("!", 1)
    return sheet_name, cell


def apply_steps(workbook_path: Path, output_path: Path, steps: list[dict]) -> None:
    wb = load_workbook(workbook_path)

    for step in steps:
        if "cell" not in step or "value" not in step:
            raise ValueError(
                f"Each fixture step must include 'cell' and 'value'. Bad step: {step}"
            )

        sheet_name, cell = parse_cell_ref(step["cell"])

        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in workbook '{workbook_path.name}'. "
                f"Available sheets: {wb.sheetnames}"
            )

        ws = wb[sheet_name]
        ws[cell] = step["value"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    if not FIXTURES_FILE.exists():
        raise FileNotFoundError(f"Fixture file not found: {FIXTURES_FILE}")

    with FIXTURES_FILE.open("r", encoding="utf-8") as f:
        fixtures = json.load(f)

    for workbook_name, steps in fixtures.items():
        source_workbook = INPUT_DIR / workbook_name
        output_workbook = OUTPUT_DIR / workbook_name

        if not source_workbook.exists():
            raise FileNotFoundError(
                f"Generated workbook not found: {source_workbook}"
            )

        if not isinstance(steps, list):
            raise ValueError(
                f"Fixture entry for '{workbook_name}' must be a list of steps."
            )

        print(f"Building test workbook: {workbook_name}")
        print(f"Source: {source_workbook}")
        print(f"Writing: {output_workbook}")

        apply_steps(source_workbook, output_workbook, steps)

        print(f"Saved: {output_workbook}")

    print("Test workbooks generated successfully.")

    
if __name__ == "__main__":
    main()