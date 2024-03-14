from .workbooklib.excel_creation_utils import (
    apply_conversion_function,
    contains_illegal_characters,
    sanitize_for_excel,
    set_range,
    sort_by_field,
)
from .exception_utils import DataMigrationValueError
from django.test import TestCase
from openpyxl import Workbook
from openpyxl.utils import quote_sheetname, absolute_coordinate
from openpyxl.workbook.defined_name import DefinedName


class ExcelCreationTests(TestCase):
    range_name = "my_range"

    def init_named_range(self, coord):
        """
        Create and return a workbook with a named range for the given coordinate
        """
        wb = Workbook()
        ws = wb.active

        # Create named range
        ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate(coord)}"
        defn = DefinedName(self.range_name, attr_text=ref)
        wb.defined_names.add(defn)

        return wb

    def test_set_range_standard(self):
        """
        Standard use case
        """
        wb = self.init_named_range("A6")
        ws = wb.active

        ws.cell(row=6, column=1, value="foo")
        self.assertEqual(ws["A6"].value, "foo")

        set_range(wb, self.range_name, ["bar"])
        self.assertEqual(ws["A6"].value, "bar")

    def test_set_range_default(self):
        """
        Using a default value
        """
        wb = self.init_named_range("A6")
        ws = wb.active

        ws.cell(row=6, column=1, value="foo")
        self.assertEqual(ws["A6"].value, "foo")

        set_range(wb, self.range_name, [None], default="bar")
        self.assertEqual(ws["A6"].value, "bar")

    def test_set_range_no_default(self):
        """
        Error when no value or default given
        """
        wb = self.init_named_range("A6")
        ws = wb.active

        ws.cell(row=6, column=1, value="foo")
        self.assertEqual(ws["A6"].value, "foo")

        self.assertRaises(
            DataMigrationValueError, set_range, wb, self.range_name, [None]
        )

    def test_set_range_conversion(self):
        """
        Applies given conversion function
        """
        wb = self.init_named_range("A6")
        ws = wb.active

        ws.cell(row=6, column=1, value="foo")
        self.assertEqual(ws["A6"].value, "foo")

        set_range(wb, self.range_name, ["1"], None, int)
        self.assertEqual(ws["A6"].value, 1)  # str -> int

    def test_set_range_multiple_values(self):
        """
        Setting multiple values
        """
        wb = self.init_named_range("A1:A2")
        ws = wb.active

        ws.cell(row=1, column=1, value=0)
        self.assertEqual(ws["A1"].value, 0)

        ws.cell(row=2, column=1, value=0)
        self.assertEqual(ws["A2"].value, 0)

        set_range(wb, self.range_name, ["1", "2"])
        self.assertEqual(ws["A1"].value, "1")
        self.assertEqual(ws["A2"].value, "2")

    def test_set_range_fewer_values(self):
        """
        Number of values is less than the range size
        """
        wb = self.init_named_range("A1:A2")
        ws = wb.active

        ws.cell(row=1, column=1, value="foo")
        self.assertEqual(ws["A1"].value, "foo")

        ws.cell(row=2, column=1, value="foo")
        self.assertEqual(ws["A2"].value, "foo")

        set_range(wb, self.range_name, ["bar"])
        self.assertEqual(ws["A1"].value, "bar")  # New value
        self.assertEqual(ws["A2"].value, "foo")  # Unchanged

    def test_set_range_multi_dests(self):
        """
        Error when multiple destinations found
        """
        wb = Workbook()
        ws = wb.active

        # Create named range with multiple destinations
        ref = f"{quote_sheetname(ws.title)}!{absolute_coordinate('A6')}, \
              {quote_sheetname(ws.title)}!{absolute_coordinate('B6')}"
        defn = DefinedName(self.range_name, attr_text=ref)
        wb.defined_names.add(defn)

        self.assertEqual(len(list(wb.defined_names[self.range_name].destinations)), 2)
        self.assertRaises(
            DataMigrationValueError, set_range, wb, self.range_name, ["bar"]
        )

    def test_set_range_ws_missing(self):
        """
        Error when the named range isn't in the given WS
        """
        wb = Workbook()

        # Create named range with bad sheet title
        ref = f"{quote_sheetname('wrong name')}!{absolute_coordinate('A6')}"
        defn = DefinedName(self.range_name, attr_text=ref)
        wb.defined_names.add(defn)

        self.assertRaises(KeyError, set_range, wb, self.range_name, ["bar"])


class TestApplyConversionFunction(TestCase):
    def test_string_conversion(self):
        """Test that a string is returned unchanged"""
        self.assertEqual(apply_conversion_function("test", "default", str), "test")

    def test_int_conversion(self):
        """Test that an int is properly returned"""
        self.assertEqual(apply_conversion_function("123", "default", int), 123)

    def test_custom_conversion(self):
        """Test that custom conversion function is applied when present"""
        self.assertEqual(
            apply_conversion_function("test", "default", lambda x: x.upper()), "TEST"
        )

    def test_default_value(self):
        """Test that a default value is returned when the input is None"""
        self.assertEqual(apply_conversion_function(None, "default", str), "default")

    def test_none_with_no_default(self):
        """Test that an exception is raised when the input is None and no default is provided"""
        with self.assertRaises(DataMigrationValueError):
            apply_conversion_function(None, None, str)


class TestExcelSanitization(TestCase):
    def test_contains_illegal_characters(self):
        """Test that contains_illegal_characters returns expected boolean values"""
        self.assertTrue(
            contains_illegal_characters("Some\x01text\x02with\x03control\x04characters")
        )  # Contains control characters
        self.assertFalse(
            contains_illegal_characters("Some text with no control characters")
        )  # No control characters
        self.assertFalse(
            contains_illegal_characters("\nNew Line\n")
        )  # Newline character is allowed

    def test_sanitize_for_excel(self):
        """Test that sanitize_for_excel returns expected values"""
        self.assertEqual(
            sanitize_for_excel("Some\x01Text\x02With\x03Control\x04Characters"),
            "SomeTextWithControlCharacters",
        )  # Control character removed
        self.assertEqual(
            sanitize_for_excel("Some text with no control characters"),
            "Some text with no control characters",
        )  # No change needed
        self.assertEqual(
            sanitize_for_excel("\nNew Line\n"), "\nNew Line\n"
        )  # Newline preserved


class TestSortRecordsByField(TestCase):
    class MockRecord:
        def __init__(self, **kwargs):
            for key, value in kwargs.items():
                setattr(self, key, value)

    def test_empty_list(self):
        """Test sorting with an empty list."""
        self.assertEqual(sort_by_field([], "some_field"), [])

    def test_sorting(self):
        """Test sorting with a non-empty list."""
        records = [
            self.MockRecord(seq_number="1"),
            self.MockRecord(seq_number="10"),
            self.MockRecord(seq_number="2"),
        ]
        sorted_records = sort_by_field(records, "seq_number")
        self.assertEqual(
            [record.seq_number for record in sorted_records], ["1", "2", "10"]
        )

    def test_sorting_with_empty_field(self):
        """Test sorting with empty sort field."""
        records = [
            self.MockRecord(other="1", seq_number=""),
            self.MockRecord(other="10", seq_number=""),
            self.MockRecord(other="2", seq_number=""),
        ]
        sorted_records = sort_by_field(records, "seq_number")
        self.assertEqual([record.other for record in sorted_records], ["1", "10", "2"])
