from datetime import datetime
import io
import logging
import uuid
import time
import openpyxl as pyxl

from boto3 import client as boto3_client
from botocore.client import ClientError, Config

from django.conf import settings
from fs.memoryfs import MemoryFS

from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname

from dissemination.models import (
    AdditionalEin,
    AdditionalUei,
    CapText,
    FederalAward,
    Finding,
    FindingText,
    General,
    Note,
    Passthrough,
    SecondaryAuditor,
    DisseminationCombined,
)

logger = logging.getLogger(__name__)

models = [
    AdditionalEin,
    AdditionalUei,
    CapText,
    FederalAward,
    Finding,
    FindingText,
    General,
    Note,
    Passthrough,
    SecondaryAuditor,
]


columns = [
    "A",
    "B",
    "C",
    "D",
    "E",
    "F",
    "G",
    "H",
    "I",
    "J",
    "K",
    "L",
    "M",
    "N",
    "O",
    "P",
    "Q",
    "R",
    "S",
    "T",
    "U",
    "V",
    "W",
    "X",
    "Y",
    "Z",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
    "AH",
    "AI",
    "AJ",
    "AK",
    "AL",
    "AM",
    "AN",
    "AO",
    "AP",
    "AQ",
    "AR",
    "AS",
    "AT",
    "AU",
    "AV",
    "AW",
    "AX",
    "AY",
    "AZ",
    "BA",
    "BB",
    "BC",
    "BD",
    "BE",
    "BF",
    "BG",
    "BH",
    "BI",
    "BJ",
]

# Each field here represents a field from the dissemination models, with the exeption of underscored fields.
# Underscored fields are generated from other disseminated data, and should be ommitted from the pre-certification reports.
# These fields are handled in "_get_attribute_or_data()"
# If one would like to add more fields that require preprocessing (such as "_aln"), ensure that they begin with an underscore.
field_name_ordered = {
    "general": [
        "report_id",
        "audit_year",
        "total_amount_expended",
        "entity_type",
        "fy_start_date",
        "fy_end_date",
        "audit_type",
        "audit_period_covered",
        "number_months",
        "auditee_uei",
        "auditee_ein",
        "auditee_name",
        "auditee_address_line_1",
        "auditee_city",
        "auditee_state",
        "auditee_zip",
        "auditee_contact_name",
        "auditee_contact_title",
        "auditee_phone",
        "auditee_email",
        "auditee_certified_date",
        "auditee_certify_name",
        "auditee_certify_title",
        "auditor_ein",
        "auditor_firm_name",
        "auditor_address_line_1",
        "auditor_city",
        "auditor_state",
        "auditor_zip",
        "auditor_country",
        "auditor_contact_name",
        "auditor_contact_title",
        "auditor_phone",
        "auditor_email",
        "auditor_foreign_address",
        "auditor_certified_date",
        "cognizant_agency",
        "oversight_agency",
        "type_audit_code",
        "sp_framework_basis",
        "is_sp_framework_required",
        "is_going_concern_included",
        "is_internal_control_deficiency_disclosed",
        "is_internal_control_material_weakness_disclosed",
        "is_material_noncompliance_disclosed",
        "gaap_results",
        "is_aicpa_audit_guide_included",
        "sp_framework_opinions",
        "agencies_with_prior_findings",
        "dollar_threshold",
        "is_low_risk_auditee",
        "is_additional_ueis",
        "date_created",
        "fac_accepted_date",
        "ready_for_certification_date",
        "submitted_date",
        "data_source",
        "is_public",
    ],
    "federalaward": [
        "report_id",
        "award_reference",
        "federal_agency_prefix",
        "federal_award_extension",
        "aln",
        "findings_count",
        "additional_award_identification",
        "federal_program_name",
        "amount_expended",
        "federal_program_total",
        "cluster_name",
        "state_cluster_name",
        "other_cluster_name",
        "cluster_total",
        "is_direct",
        "is_passthrough_award",
        "passthrough_amount",
        "is_major",
        "audit_report_type",
        "is_loan",
        "loan_balance",
    ],
    "finding": [
        "report_id",
        "federal_agency_prefix",
        "federal_award_extension",
        "aln",
        "award_reference",
        "reference_number",
        "type_requirement",
        "is_modified_opinion",
        "is_other_findings",
        "is_material_weakness",
        "is_significant_deficiency",
        "is_other_matters",
        "is_questioned_costs",
        "is_repeat_finding",
        "prior_finding_ref_numbers",
    ],
    "findingtext": [
        "id",
        "report_id",
        "finding_ref_number",
        "contains_chart_or_table",
        "finding_text",
    ],
    "note": [
        "id",
        "report_id",
        "note_title",
        "is_minimis_rate_used",
        "accounting_policies",
        "rate_explained",
        "content",
        "contains_chart_or_table",
    ],
    "captext": [
        "report_id",
        "finding_ref_number",
        "planned_action",
        "contains_chart_or_table",
    ],
    "additionalein": ["report_id", "additional_ein"],
    "additionaluei": ["report_id", "additional_uei"],
    "passthrough": [
        "report_id",
        "award_reference",
        "passthrough_name",
        "passthrough_id",
    ],
    "secondaryauditor": [
        "report_id",
        "auditor_name",
        "auditor_ein",
        "address_street",
        "address_city",
        "address_state",
        "address_zipcode",
        "contact_name",
        "contact_title",
        "contact_email",
        "contact_phone",
    ],
}

restricted_model_names = ["captext", "findingtext", "note"]

limit_disclaimer = f"This spreadsheet contains the first {settings.SUMMARY_REPORT_DOWNLOAD_LIMIT} results of your search. If you need to download more than {settings.SUMMARY_REPORT_DOWNLOAD_LIMIT} submissions, try limiting your search parameters to download in batches."
can_read_tribal_disclaimer = "This document includes one or more Tribal entities that have chosen to keep their data private per 2 CFR 200.512(b)(2). Because your account has access to these submissions, this document includes their audit findings text, corrective action plan, and notes to SEFA. Don't share this data outside your agency."
cannot_read_tribal_disclaimer = "This document includes one or more Tribal entities that have chosen to keep their data private per 2 CFR 200.512(b)(2). It doesn't include their audit findings text, corrective action plan, or notes to SEFA."


def _get_model_by_name(name):
    for m in models:
        if m.__name__.lower() == name:
            return m
    return None


def get_tribal_report_ids(report_ids):
    t0 = time.time()
    """Filters the given report_ids with only ones that are tribal"""
    objects = General.objects.all().filter(report_id__in=report_ids, is_public=False)
    objs = [obj.report_id for obj in objects]
    t1 = time.time()
    return (objs, t1 - t0)


def set_column_widths(worksheet):
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max(
                    (dims.get(cell.column, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        # Pad the column by a bit, so things are not cramped.
        worksheet.column_dimensions[columns[col - 1]].width = int(value * 1.2)


def protect_sheet(sheet):
    sheet.protection.sheet = True
    sheet.protection.password = str(uuid.uuid4())
    sheet.protection.enable()


def insert_precert_coversheet(workbook):
    sheet = workbook.create_sheet("Coversheet", 0)
    sheet.append(["Time created", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
    sheet.append(["Note", "This file is for review only and can't be edited."])
    set_column_widths(sheet)
    protect_sheet(sheet)


def insert_dissem_coversheet(workbook, contains_tribal, include_private):
    sheet = workbook.create_sheet("Coversheet", 0)
    sheet.append(["Time created", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")])
    sheet.append(
        [
            "Note",
            limit_disclaimer,
        ]
    )

    if contains_tribal:
        if include_private:
            sheet.append(
                [
                    "Note",
                    can_read_tribal_disclaimer,
                ]
            )
        else:
            sheet.append(
                [
                    "Note",
                    cannot_read_tribal_disclaimer,
                ]
            )

    # Uncomment if we want to link to the FAC API for larger data dumps.
    # sheet.cell(row=3, column=2).value = "FAC API Link"
    # sheet.cell(row=3, column=2).hyperlink = f"{settings.STATIC_SITE_URL}/developers/"
    set_column_widths(sheet)


def _get_attribute_or_data(obj, field_name):
    value = getattr(obj, field_name)
    if isinstance(value, General):
        value = value.report_id
    return value


def gather_report_data_dissemination(report_ids, tribal_report_ids, include_private):
    """
    Given a set of report IDs, fetch the disseminated data for each and asssemble into a dictionary with the following shape:

    {
        general: {
            field_names: [],
            entries: [],
        },
        federal_award: {
            field_names: [],
            entries: [],
        },
        ...
    }
    """
    t0 = time.time()
    # Make report IDs unique
    report_ids = set(report_ids)
    all_names = set(field_name_ordered.keys())
    names_in_dc = set(["general", "federalaward", "finding", "passthrough"])
    names_not_in_dc = all_names - names_in_dc
    data = initialize_data_structure(names_in_dc.union(names_not_in_dc))

    process_combined_results(
        report_ids, names_in_dc, data, include_private, tribal_report_ids
    )

    process_non_combined_results(
        report_ids, names_not_in_dc, data, include_private, tribal_report_ids
    )

    return (data, time.time() - t0)


def initialize_data_structure(names):
    data = {}
    for model_name in names:
        data[model_name] = {
            "field_names": field_name_ordered[model_name],
            "entries": [],
        }
    return data


def process_combined_results(
    report_ids, names_in_dc, data, include_private, tribal_report_ids
):
    # Grab all the rows from the combined table into a local structure.
    # We'll do this in memory. This table flattens general, federalaward, and findings
    # so we can move much faster on those tables without extra lookups.
    dc_results = DisseminationCombined.objects.all().filter(report_id__in=report_ids)
    # Different tables want to be visited/filtered differently.
    visited = set()
    # Do all of the names in the DisseminationCombined at the same time.
    # That way, we only go through the results once.
    for obj in dc_results:
        for model_name in names_in_dc:
            field_names = field_name_ordered[model_name]
            report_id = getattr(obj, "report_id")
            award_reference = getattr(obj, "award_reference")
            reference_number = getattr(obj, "reference_number")
            passthrough_name = getattr(obj, "passthrough_name")

            # WATCH THIS IF/ELIF
            # It is making sure we do not double-disseminate some rows.
            ####
            # GENERAL
            if model_name == "general" and report_id in visited:
                pass
            ####
            # PASSTHROUGH
            # We should never disseminate something that has no name.
            elif model_name == "passthrough" and passthrough_name is None:
                pass
            ####
            # FEDERAL AWARD
            # This condition is actually filtering out the damage to the
            # data from the race hazard we had at the start of 2024.
            # NOTE
            # We cannot filter `passthrough` here. Each award reference row has
            # a one-to-many relationship with passthrough.
            elif (
                model_name == "federalaward"
                and f"{report_id}-{award_reference}" in visited
            ):
                pass
            ####
            # FINDING
            elif model_name == "finding" and (
                award_reference is None or reference_number is None
            ):
                # And we don't include rows in finding where there are none.
                pass
            else:
                # Track to limit duplication
                if model_name == "general":
                    visited.add(report_id)
                # Handle special tracking for federal awards, so we don't duplicate award # rows.
                if model_name == "federalaward":
                    visited.add(f"{report_id}-{award_reference}")
                # Omit rows for private tribal data when the user doesn't have perms
                if (
                    model_name in restricted_model_names
                    and not include_private
                    and report_id in tribal_report_ids
                ):
                    pass
                else:
                    data[model_name]["entries"].append(
                        [getattr(obj, field_name) for field_name in field_names]
                    )


def process_non_combined_results(
    report_ids, names_not_in_dc, data, include_private, tribal_report_ids
):
    for model_name in names_not_in_dc:
        model = _get_model_by_name(model_name)
        print(model_name)
        field_names = field_name_ordered[model_name]
        objects = model.objects.all().filter(report_id__in=report_ids)
        # Walk the objects
        for obj in objects:
            report_id = _get_attribute_or_data(obj, "report_id")
            # Omit rows for private tribal data when the user doesn't have perms
            if (
                model_name in restricted_model_names
                and not include_private
                and report_id in tribal_report_ids
            ):
                pass
            else:
                data[model_name]["entries"].append(
                    [
                        _get_attribute_or_data(obj, field_name)
                        for field_name in field_names
                    ]
                )


def gather_report_data_pre_certification(i2d_data):
    """
    Given a sac object dict, asssemble into a dictionary with the following shape:

    {
        general: {
            field_names: [],
            entries: [],
        },
        federal_award: {
            field_names: [],
            entries: [],
        },
        ...
    }
    """

    # Map IntakeToDissemination names to the dissemination table names
    i2d_to_dissemination = {
        "Generals": General,
        "SecondaryAuditors": SecondaryAuditor,
        "FederalAwards": FederalAward,
        "Findings": Finding,
        "FindingTexts": FindingText,
        "Passthroughs": Passthrough,
        "CapTexts": CapText,
        "Notes": Note,
        "AdditionalUEIs": AdditionalUei,
        "AdditionalEINs": AdditionalEin,
    }

    # Move the IntakeToDissemination data to dissemination_data, under the proper naming scheme.
    dissemination_data = {}
    for name_i2d, model in i2d_to_dissemination.items():
        dissemination_data[model.__name__.lower()] = i2d_data.get(name_i2d, [])

    data = {}

    # FIXME
    # This is because additional fields were added to the SF-SAC.
    # Then, we introduced DisseminationCombined
    # Then, we got rid of `_` on field names.
    # Now, this broke.
    # Choices were made, consequences followed. We want to clean this up.
    fields_to_ignore = {
        "federalaward": ["aln"],
        "finding": ["aln", "federal_agency_prefix", "federal_award_extension"],
    }

    # For every model (FederalAward, CapText, etc), add the skeleton object ('field_names' and empty 'entries') to 'data'.
    # Then, for every object in the dissemination_data (objects of type FederalAward, CapText, etc) create a row (array) for the summary.
    # Every row is created by looping over the field names and appending the data.
    # We also strip tzinfo from the dates, because excel doesn't like them.
    # Once a row is created, append it to the data[ModelName]['entries'] array.
    for model in models:
        model_name = (
            model.__name__.lower()
        )  # Sheet/tab name, ex. "federalaward", "finding"
        field_names = [
            field_name
            for field_name in field_name_ordered[model_name]
            if field_name not in fields_to_ignore.get(model_name, [])
        ]  # Column names, omitting "_" fields
        data[model_name] = {
            "field_names": field_names,
            "entries": [],
        }  # The sheet/tab content, with no rows by default

        # For every instance of a model we have, generate an appropriate row and append it to the sheet/tab.
        # Ignore underscored fields and wipe timezone information.
        for obj in dissemination_data[model_name]:
            row = []
            for field_name in field_names:
                if not field_name.startswith("_"):
                    value = getattr(obj, field_name)
                    if isinstance(value, datetime):
                        value = value.replace(tzinfo=None)
                    if isinstance(value, General):
                        value = value.report_id
                    row.append(value)
            data[model_name]["entries"].append(row)

    return data


def create_workbook(data, protect_sheets=False):
    t0 = time.time()
    workbook = pyxl.Workbook()

    for sheet_name in data.keys():
        sheet = workbook.create_sheet(sheet_name)

        # create a header row with the field names
        sheet.append(data[sheet_name]["field_names"])

        # append a new row for each entry in the dataset
        for entry in data[sheet_name]["entries"]:
            sheet.append(entry)

        # add named ranges for the columns, now that the data is loaded.
        for index, field_name in enumerate(data[sheet_name]["field_names"]):
            coordinate = f"${columns[index]}$2:${columns[index]}${2 + len(data[sheet_name]['entries'])}"
            ref = f"{quote_sheetname(sheet.title)}!{coordinate}"
            named_range = DefinedName(f"{sheet_name}_{field_name}", attr_text=ref)
            workbook.defined_names.add(named_range)

        set_column_widths(sheet)
        if protect_sheets:
            protect_sheet(sheet)

    # remove sheet that is created during workbook construction
    workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))
    t1 = time.time()
    return (workbook, t1 - t0)


def persist_workbook(workbook):
    t0 = time.time()
    s3_client = boto3_client(
        service_name="s3",
        region_name=settings.AWS_S3_PRIVATE_REGION_NAME,
        aws_access_key_id=settings.AWS_PRIVATE_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_PRIVATE_SECRET_ACCESS_KEY,
        endpoint_url=settings.AWS_S3_PRIVATE_INTERNAL_ENDPOINT,
        config=Config(signature_version="s3v4"),
    )

    with MemoryFS() as mem_fs:
        now = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        filename = f"fac-summary-report-{now}.xlsx"
        s3_dir = "temp"

        workbook_write_fp = mem_fs.openbin(filename, mode="w")
        workbook.save(workbook_write_fp)
        workbook_read_fp = mem_fs.openbin(filename, mode="r")
        workbook_read_fp.seek(0)
        content = workbook_read_fp.read()
        workbook_bytes = io.BytesIO(content)

        try:
            s3_client.put_object(
                Body=workbook_bytes,
                Bucket=settings.AWS_PRIVATE_STORAGE_BUCKET_NAME,
                Key=f"{s3_dir}/{filename}",
            )
        except ClientError:
            logger.warn(f"Unable to put summary report file {filename} in S3!")
            raise
    t1 = time.time()
    return (f"temp/{filename}", t1 - t0)


def generate_summary_report(report_ids, include_private=False):
    t0 = time.time()
    (tribal_report_ids, ttri) = get_tribal_report_ids(report_ids)
    (data, tgrdd) = gather_report_data_dissemination(
        report_ids, tribal_report_ids, include_private
    )
    (workbook, tcw) = create_workbook(data)
    insert_dissem_coversheet(workbook, bool(tribal_report_ids), include_private)
    (filename, tpw) = persist_workbook(workbook)
    t1 = time.time()
    logger.info(
        f"SUMMARY_REPORTS generate_summary_report\n\ttotal: {t1-t0} ttri: {ttri} tgrdd: {tgrdd} tcw: {tcw} tpw: {tpw}"
    )
    return filename


# Ignore performance profiling for the presub.
def generate_presubmission_report(i2d_data):
    data = gather_report_data_pre_certification(i2d_data)
    (workbook, _) = create_workbook(data, protect_sheets=True)
    insert_precert_coversheet(workbook)
    workbook.security.workbookPassword = str(uuid.uuid4())
    workbook.security.lockStructure = True
    (filename, _) = persist_workbook(workbook)

    return filename
