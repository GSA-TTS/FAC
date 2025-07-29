from datetime import datetime, timezone
from dissemination.models import General, Finding, FederalAward
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io
import time
import logging
from django.forms.models import model_to_dict

logger = logging.getLogger(__name__)


def lmap(fun, ls):
    return list(map(fun, ls))


def objs_to_dict(lob):
    return lmap(model_to_dict, lob)


class Timing:
    def __init__(self, tag):
        self.tag = tag

    def __enter__(self):
        self.start = time.time()

    def __exit__(self, exception_type, exception_value, traceback):
        self.end = time.time()
        logger.info(f"{self.tag}: {self.end - self.start}")


def get_award(awards, report_id, award_reference):
    key = tuple([report_id, award_reference])
    if key in awards:
        return awards[key]
    else:
        return None


def get_findings_for_agency_number(report_ids):
    with Timing("FINDINGS GFFAN PREP"):
        # findings_for_report_ids takes a set of report ids and gives us
        # the findings objects associated with those reports
        with Timing("FINDINGS GFFAN FFRI"):
            findings_for_report_ids = lmap(
                model_to_dict, Finding.objects.filter(report_id__in=report_ids)
            )
        # We can now get a list of report ids that have findings. This will
        # be a subset of `report_ids`
        with Timing("FINDINGS GFFAN RIWF"):
            report_ids_with_findings = set(
                lmap(lambda d: d["report_id"], findings_for_report_ids)
            )

        # The awards rows with findings will have a report id in the subset.
        # We are also doing this agency-number-by-agency-number, so we'll further restrict things.
        with Timing("FINDINGS GFFAN AWF"):
            awards_with_findings = lmap(
                model_to_dict,
                FederalAward.objects.filter(
                    report_id__in=report_ids_with_findings, findings_count__gte=1
                ),
            )

        # To speed things up, key those by report-id/award-ref
        # This builds a lookup table to directly find awards associated with a report.
        with Timing("FINDINGS GFFAN AWFK"):
            awards_with_findings_keyed = {}
            for a in awards_with_findings:
                key = tuple([a["report_id"], a["award_reference"]])
                awards_with_findings_keyed[key] = a

        # These are all of the general objects that have findings. Again, we're subsetting.
        gobjs = lmap(
            model_to_dict,
            General.objects.filter(report_id__in=report_ids_with_findings),
        )

    uniq = set()
    results = {}

    # I want to start by going report-by-report.
    with Timing("FINDINGS GUAN"):
        for gobj in gobjs:
            # For a given report, I know there are findings. I want unique findings rows. So, I'll
            # go through the findings that have this report ID.
            for fobj in filter(
                lambda d: d["report_id"] == gobj["report_id"], findings_for_report_ids
            ):
                # I now want to output one row per unique (report_id, award_ref, ref_num).
                key = tuple(
                    [
                        gobj["report_id"],
                        fobj["award_reference"],
                        fobj["reference_number"],
                    ]
                )
                if key not in uniq:
                    # This makes sure I only do this once.
                    uniq.add(key)
                    # Now, I need an award that matches this report id and reference.
                    aobj = get_award(
                        awards_with_findings_keyed,
                        gobj["report_id"],
                        fobj["award_reference"],
                    )
                    # The awards are constrained to agency number; the findings are not.
                    # We might not find an award, meaning we need to skip for this tab in the sheet.
                    if aobj is not None:
                        d = {}
                        d["report_id"] = gobj["report_id"]
                        d["auditee_name"] = gobj["auditee_name"]
                        d["auditee_uei"] = gobj["auditee_uei"]
                        d["award_reference"] = fobj["award_reference"]
                        d["reference_number"] = fobj["reference_number"]
                        d["aln"] = (
                            aobj["federal_agency_prefix"]
                            + "."
                            + aobj["federal_award_extension"]
                        )
                        d["cog_over"] = "COG" if gobj["cognizant_agency"] else "OVER"
                        d["cog_over_agency"] = (
                            f"{gobj['cognizant_agency']}"
                            if gobj["cognizant_agency"]
                            else f"{gobj['oversight_agency']}"
                        )
                        d["federal_program_name"] = aobj["federal_program_name"]
                        d["amount_expended"] = aobj["amount_expended"]
                        d["is_direct"] = aobj["is_direct"]
                        d["is_major"] = aobj["is_major"]
                        d["is_passthrough_award"] = aobj["is_passthrough_award"]
                        d["passthrough_amount"] = aobj["passthrough_amount"]
                        d["is_modified_opinion"] = fobj["is_modified_opinion"]
                        d["is_other_matters"] = fobj["is_other_matters"]
                        d["is_material_weakness"] = fobj["is_material_weakness"]
                        d["is_significant_deficiency"] = fobj[
                            "is_significant_deficiency"
                        ]
                        d["is_other_findings"] = fobj["is_other_findings"]
                        d["is_questioned_costs"] = fobj["is_questioned_costs"]
                        d["is_repeat_finding"] = fobj["is_repeat_finding"]
                        d["prior_finding_ref_numbers"] = fobj[
                            "prior_finding_ref_numbers"
                        ]
                        # This builds a dictionary based on agency prefix.
                        # Each agency number becomes a tab in the spreadsheet.
                        if aobj["federal_agency_prefix"] in results:
                            results[aobj["federal_agency_prefix"]].append(d)
                        else:
                            results[aobj["federal_agency_prefix"]] = [d]
    return results


def adjust_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            if cell and cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    return ws


def add_sheets(wb, iter, query):
    # iter_value will  be an integer
    for iter_value in iter:
        res = query(iter_value)
        if len(res) > 0:
            ws = wb.create_sheet(f"{iter_value}")
            # Put headers on the sheets
            for d in res:
                ws.append(list(d.keys()))
                break
            # Now the values.
            for d in res:
                ws.append(list(d.values()))
            adjust_columns(ws)


def convert_bools(res):
    for k in res.keys():
        if res[k] in ["Y", "TRUE", "T", "YES"]:
            res[k] = True
        elif res[k] in ["N", "NO", "FALSE", "F"]:
            res[k] = False
    return res


yes_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")


def cleanup_report_ids(ws):
    report_ids = []
    if ws:
        for cell in ws["A"]:
            if ("GSAFAC" in str(cell.value)) or ("CENSUS" in str(cell.value)):
                report_ids.append(cell.value)
                cell.hyperlink = (
                    f"https://app.fac.gov/dissemination/report/pdf/{cell.value}"
                )
            else:
                pass
        return report_ids


def cleanup_summary_links(ws, report_ids):
    if ws:
        for ndx, cell in enumerate(ws["B"][1:]):
            if cell:
                cell.hyperlink = (
                    f"https://app.fac.gov/dissemination/summary/{report_ids[ndx]}"
                )


def cleanup_booleans(ws):
    boolean_columns = ["K", "L", "M", "O", "P", "Q", "R", "S", "T", "U"]
    for bool_column in boolean_columns:
        for cell in ws[bool_column]:
            if cell.value in [1, "Y"]:
                cell.value = "YES"
            elif cell.value in [0, "N"]:
                cell.value = "NO"
            else:
                pass
    for bool_column in boolean_columns:
        for cell in ws[bool_column]:
            if cell.value == "YES":
                cell.fill = yes_fill


def cleanup_sheet(ws):
    # Trys to go through a sheet and
    # 1. Hyperlink all the report ids,
    # 2. Cleanup all the booleans.
    # The columns are hard-coded to the order
    # they appear from the dump into the sheet.
    report_ids = cleanup_report_ids(ws)
    cleanup_summary_links(ws, report_ids)
    cleanup_booleans(ws)


def remove_default_sheet(wb):
    # Try removing the default sheet.
    try:
        del wb["Sheet"]
    except KeyError:
        pass


def gather_report_data(report_ids=[]):
    # Now, lets start building our workbook. It's an iterative
    # set of queries.
    wb = Workbook()
    with Timing("FINDINGS DATA PREP"):
        results = get_findings_for_agency_number(report_ids)

    with Timing("FINDINGS ADD SHEETS"):
        add_sheets(
            wb,
            sorted(results.keys()),
            lambda an: sorted(
                results[an],
                key=lambda d: f"{d['report_id']}-{d['award_reference']}-{d['reference_number']}",
            ),
        )

    # Hyperlink the report IDs
    for sheet in wb.worksheets:
        cleanup_sheet(sheet)

    remove_default_sheet(wb)

    return wb


def prepare_workbook_for_download(workbook):
    now = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    workbook_filename = f"findings-summary-{now}.xlsx"

    # Save the workbook directly to a BytesIO object
    workbook_bytes = io.BytesIO()
    workbook.save(workbook_bytes)
    workbook_bytes.seek(0)

    return workbook_filename, workbook_bytes


# This generates reports of the form that we generated
# while advanced search was offline.
#
# These reports *only* contain public data, and therefore
# do not have access control concerns.
def generate_findings_summary_report(report_ids=[]):
    with Timing("FINDINGS TOTAL"):
        with Timing("FINDINGS GATHER DATA"):
            wb = gather_report_data(report_ids)
        with Timing("FINDINGS PREP WORKBOOK DOWNLOAD"):
            (filename, workbook_bytes) = prepare_workbook_for_download(wb)
    return filename, workbook_bytes
