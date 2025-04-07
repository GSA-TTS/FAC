from unittest import TestCase


from model_bakery import baker
import openpyxl as pyxl

from audit.fixtures.audit_information import fake_audit_information
from audit.fixtures.certification import (
    fake_auditee_certification,
    fake_auditor_certification,
)

from audit.models import Audit
from audit.models.constants import ORGANIZATION_TYPE, STATUS
from audit.test_views import (
    build_auditee_cert_dict,
    build_auditor_cert_dict,
    load_json,
    AUDIT_JSON_FIXTURES,
)
from dissemination.report_generation.audit_summary_reports import (
    generate_audit_summary_report,
    _tribal_report_ids,
    _gather_audit_data,
)
from dissemination.report_generation.excel.coversheets import (
    insert_dissemination_coversheet,
    can_read_tribal_disclaimer,
    cannot_read_tribal_disclaimer,
)


class SummaryReportTests(TestCase):
    def test_generate_summary_report_returns_filename(self):
        """The filename returned should be correctly formatted"""
        audit = baker.make(Audit, version=0, _quantity=100)
        report_ids = [a.report_id for a in audit]

        filename, _ = generate_audit_summary_report(report_ids)

        self.assertTrue(filename.startswith, "fac-summary-report-")
        self.assertTrue(filename.endswith, ".xlsx")

    def test_contains_private_tribal(self):
        """The report_ids returned should only belong to tribal audits"""
        public_general = baker.make(
            Audit,
            version=0,
            audit={
                "is_public": True,
                "general_information": {
                    "user_provided_organization_type": ORGANIZATION_TYPE.STATE
                },
            },
            _quantity=3,
        )
        tribal_general = baker.make(
            Audit,
            version=0,
            audit={
                "is_public": False,
                "general_information": {
                    "user_provided_organization_type": ORGANIZATION_TYPE.TRIBAL
                },
            },
            _quantity=2,
        )

        public_report_ids = [g.report_id for g in public_general]
        tribal_report_ids = [g.report_id for g in tribal_general]

        tribal_report_ids, _ = _tribal_report_ids(public_report_ids + tribal_report_ids)
        self.assertEqual(len(tribal_report_ids), 2)

    def test_contains_private_tribal_no_tribal(self):
        """No report_ids should be returned for entirely public audits"""
        public_general = baker.make(
            Audit, version=0, audit={"is_public": True}, _quantity=3
        )
        public_report_ids = [g.report_id for g in public_general]

        tribal_report_ids, _ = _tribal_report_ids(public_report_ids)
        self.assertEqual(len(tribal_report_ids), 0)

    def test_contains_public_tribal(self):
        """The report_ids returned should only belong to tribal audits"""
        public_general = baker.make(
            Audit,
            version=0,
            audit={
                "is_public": True,
                "general_information": {
                    "user_provided_organization_type": ORGANIZATION_TYPE.STATE
                },
            },
            _quantity=3,
        )
        tribal_general = baker.make(
            Audit,
            version=0,
            audit={
                "is_public": True,
                "general_information": {
                    "user_provided_organization_type": ORGANIZATION_TYPE.TRIBAL
                },
            },
            _quantity=2,
        )

        public_report_ids = [g.report_id for g in public_general]
        tribal_report_ids = [g.report_id for g in tribal_general]

        tribal_ids, _ = _tribal_report_ids(public_report_ids + tribal_report_ids)
        self.assertEqual(len(tribal_ids), 0)

    def test_insert_dissem_coversheet_tribal_can_read(self):
        """Coversheet should have correct disclaimer for tribal data that can be read"""
        workbook = pyxl.Workbook()
        insert_dissemination_coversheet(workbook, True, True)
        coversheet = workbook.get_sheet_by_name("Coversheet")

        self.assertEqual(coversheet["B3"].value, can_read_tribal_disclaimer)

    def test_insert_dissem_coversheet_tribal_cannot_read(self):
        """Coversheet should have correct disclaimer for tribal data that cannot be read"""
        workbook = pyxl.Workbook()
        insert_dissemination_coversheet(workbook, True, False)
        coversheet = workbook.get_sheet_by_name("Coversheet")

        self.assertEqual(coversheet["B3"].value, cannot_read_tribal_disclaimer)

    def test_insert_dissem_coversheet_public(self):
        """Coversheet should have no disclaimer for tribal data if all data is public"""
        workbook = pyxl.Workbook()
        insert_dissemination_coversheet(workbook, False, False)
        coversheet = workbook.get_sheet_by_name("Coversheet")

        self.assertEqual(coversheet["B3"].value, None)

    def _test_gather_report_data_dissemination_helper(self, include_private):
        """
        Helper to create public and tribal audits, generate a summary, and check the
        correct tribal data is omitted
        """
        # Create public audit with relevant sections
        public = _create_audit(False, True)
        public_report_ids = [public.report_id]

        # Create tribal audit with relevant sections
        tribal = _create_audit(True, False)
        tribal_report_ids = [tribal.report_id]

        # Get the data that constitutes the summary workbook
        (data, _) = _gather_audit_data(
            public_report_ids + tribal_report_ids,
            tribal_report_ids,
            include_private,
        )
        # These sheets should all omit rows containing the tribal report_id when !include_private
        private_sheets = ["captext", "note", "findingtext"]
        for sheet in private_sheets:
            found_public, found_tribal = False, False
            report_id_index = data[sheet]["field_names"].index("report_id")
            for row in data[sheet]["entries"]:
                if row[report_id_index] == public.report_id:
                    found_public = True
                elif row[report_id_index] == tribal.report_id:
                    found_tribal = True

            self.assertTrue(found_public)
            self.assertEqual(found_tribal, include_private)

    def test_gather_report_data_dissemination_exclude_private(self):
        """Summaries with tribal data and no access"""
        self._test_gather_report_data_dissemination_helper(False)

    def test_gather_report_data_dissemination_include_private(self):
        """Summaries with tribal data and access"""
        self._test_gather_report_data_dissemination_helper(True)


def _create_audit(is_tribal, is_public):
    geninfofile = "general-information--test0001test--simple-pass.json"
    audit_data = {
        "is_public": is_public,
        "auditee_certification": build_auditee_cert_dict(*fake_auditee_certification()),
        "auditor_certification": build_auditor_cert_dict(*fake_auditor_certification()),
        "audit_information": fake_audit_information(),
        "general_information": load_json(AUDIT_JSON_FIXTURES / geninfofile),
        "notes_to_sefa": {
            "accounting_policies": "Exhaustive",
            "is_minimis_rate_used": "Y",
            "rate_explained": "At great length",
            "notes_to_sefa_entries": [
                {
                    "note_title": "1. Title",
                    "seq_number": 1,
                    "note_content": "Content",
                    "contains_chart_or_table": "N",
                }
            ],
        },
        "findings_text": [
            {
                "text_of_finding": "Finding",
                "reference_number": "2023-002",
                "contains_chart_or_table": "N",
            }
        ],
        "findings_uniform_guidance": [
            {
                "program": {
                    "award_reference": "AWARD-00001",
                    "compliance_requirement": "AB",
                },
                "findings": {
                    "is_valid": "Y",
                    "prior_references": "N/A",
                    "reference_number": "2023-002",
                    "repeat_prior_reference": "N",
                },
                "other_matters": "N",
                "other_findings": "N",
                "modified_opinion": "N",
                "questioned_costs": "Y",
                "material_weakness": "N",
                "significant_deficiency": "Y",
            }
        ],
        "corrective_action_plan": [
            {
                "planned_action": "Fix it.",
                "reference_number": "2023-002",
                "contains_chart_or_table": "N",
            }
        ],
        "federal_awards": {
            "total_amount_expended": 1_000_000,
            "awards": [
                {
                    "cluster": {"cluster_name": "N/A", "cluster_total": 0},
                    "program": {
                        "is_major": "N",
                        "program_name": "Name",
                        "amount_expended": 1_000_000,
                        "federal_agency_prefix": "98",
                        "federal_program_total": 1_000_000,
                        "three_digit_extension": "001",
                        "number_of_audit_findings": 0,
                        "additional_award_identification": "AWARD",
                    },
                    "direct_or_indirect_award": {"is_direct": "Y"},
                    "award_reference": "AWARD-00001",
                }
            ],
        },
    }
    if is_tribal:
        audit_data = audit_data | {
            "general_information": {
                "user_provided_organization_type": ORGANIZATION_TYPE.TRIBAL
            }
        }

    audit = baker.make(
        Audit, version=0, submission_status=STATUS.DISSEMINATED, audit=audit_data
    )

    return audit
