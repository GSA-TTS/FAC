from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname, absolute_coordinate
from openpyxl.styles import PatternFill, Alignment, Protection, Font

# from openpyxl.styles.colors import Color
# from openpyxl.worksheet.dimensions import ColumnDimension

import json

# from json import JSONEncoder
import _jsonnet

import os
import sys

from xkcdpass import xkcd_password as xp
from openpyxl.workbook.protection import WorkbookProtection

from collections import namedtuple as NT

Range = NT(
    "Range",
    "name,column,label_row,range_start_row,range_start,abs_range_start,full_range",
)

MAGIC_MAX_ROWS = 1048576
MAX_ROWS = 3000


# Styling
header_row_fill = PatternFill(
    fill_type="solid",
    start_color="000066CC",
    end_color="000066CC",
    bgColor="000066CC",
    patternType="solid",
)

header_row_font = Font(
    name="Calibri",
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline="none",
    strike=False,
    color="00FFFFFF",
)


def jsonnet_sheet_spec_to_json(filename):
    json_str = _jsonnet.evaluate_snippet(filename, open(filename).read())
    jobj = json.loads(json_str)
    return jobj


def create_empty_workbook():
    wb = Workbook()
    ws = wb.active
    # Remove the default/active worksheet. We'll make more.
    wb.remove(ws)
    return wb


def create_protected_sheet(wb, spec, password, ndx):
    ws = wb.create_sheet(title=spec["name"], index=ndx)
    # The sheet has to be locked, and individual cells unlocked.
    # https://stackoverflow.com/questions/46877091/lock-some-cells-from-editing-in-python-openpyxl
    ws.protection.set_password(password)
    ws.protection.sheet = True
    ws.protection.enable()
    return ws


def merge_adjacent_columns(ws, cell_ranges):
    """
    Merges cells in adjacent columns for each row within the specified range.
    """
    for cell_range in cell_ranges:
        start_row, end_row, start_column, end_column = cell_range
        for row in range(start_row, end_row):
            merge_range = f"{start_column}{row}:{end_column}{row}"
            ws.merge_cells(merge_range)


def process_single_cells(wb, ws, sheet):
    # Create all the single cells
    for o in sheet["single_cells"]:
        cell_coordinate = o["range_cell"]
        absolute_cell_coordinate = f"{absolute_coordinate(cell_coordinate)}"
        cell_row = int(o["range_cell"][1])
        cell_column = o["range_cell"][0]
        sheet_cell_coordinate = (
            f"{quote_sheetname(ws.title)}!{absolute_cell_coordinate}"
        )
        print(f"Single Cell: {absolute_cell_coordinate} {sheet_cell_coordinate}")
        new_range = DefinedName(name=o["range_name"], attr_text=sheet_cell_coordinate)
        wb.defined_names.add(new_range)
        the_cell = ws[o["title_cell"]]
        the_cell.value = o["title"]
        the_cell.fill = header_row_fill
        the_cell.font = header_row_font
        the_cell.alignment = Alignment(wrapText=True, wrap_text=True)
        # ws.row_dimensions[cell_row].height = 40
        entry_cell_obj = ws[absolute_cell_coordinate]
        entry_cell_obj.protection = Protection(locked=False)
        # Creating a coord for the single-cell range for validation configuration
        coord = Range(
            "",
            cell_column,
            0,
            cell_row,
            absolute_cell_coordinate,
            "",
            f"{o['range_cell']}:{o['range_cell']}",
        )
        configure_validation(ws, coord, o)
        # Individual cells can have width set, optionally
        if "width" in o:
            ws.column_dimensions[o["title_cell"][0]].width = o["width"]


def make_range(r):
    column = r["title_cell"][0]
    title_row = int(r["title_cell"][1])
    range_start_row = int(r["title_cell"][1]) + 1
    start_cell = column + str(range_start_row)
    full_range = f"${column}${range_start_row}:${column}${MAX_ROWS}"
    return Range(
        r["range_name"],
        column,
        title_row,
        range_start_row,
        start_cell,
        f"{absolute_coordinate(start_cell)}",
        full_range,
    )


def configure_header_cell(ws, r):
    the_cell = ws[r["title_cell"]]
    the_cell.value = r["title"]
    the_cell.fill = header_row_fill
    the_cell.font = header_row_font
    the_cell.alignment = Alignment(wrapText=True, wrap_text=True)


def apply_header_cell_style(ws, additional_header_cells):
    for ahc in additional_header_cells:
        the_cell = ws[ahc]
        the_cell.fill = header_row_fill


def process_open_ranges(wb, ws, spec):
    for r in spec["open_ranges"]:
        coords = make_range(r)
        cell_reference = f"{quote_sheetname(ws.title)}!{coords.full_range}"
        print(
            f"Open Range: {r['title']}, {coords.column}, {coords.abs_range_start}, {cell_reference}"
        )
        new_range = DefinedName(name=coords.name, attr_text=cell_reference)
        wb.defined_names.add(new_range)
        configure_header_cell(ws, r)
        # Make the header row tall.
        ws.row_dimensions[coords.range_start_row - 1].height = 100


def add_yorn_validation(ws):
    yorn_dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    yorn_dv.error = "Entries must be Y or N in this column"
    yorn_dv.errorTitle = "Must by Y or N"
    yorn_dv.showDropDown = False
    # https://stackoverflow.com/questions/75889368/openpyxl-excel-file-created-is-not-showing-validation-errors-or-prompt-message
    yorn_dv.showErrorMessage = True
    ws.add_data_validation(yorn_dv)
    return yorn_dv


def configure_validation(ws, coords: Range, r):
    dv = None
    if r["type"] == "yorn_range":
        dv = add_yorn_validation(ws)
        dv.add(coords.full_range)
    if ("validation" in r) and (r["validation"] != {}):
        v = r["validation"]
        dv = DataValidation(type=v["type"])
        if "formula1" in v:
            dv.formula1 = v["formula1"]
            dv.formula1 = dv.formula1.replace(
                "FIRSTCELLREF", f"${coords.column}{coords.range_start_row}"
            )
        # elif "enum" in v:
        #     valid_values = v["enum"]
        #     dv.formula1 = f'"IF(ISNA(MATCH({coords.column}{coords.range_start_row}, {{{",".join([f"{chr(34)}{x}{chr(34)}" for x in valid_values])}}}, 0)),FALSE,TRUE)"'
        if "operator" in v:
            dv.operator = v["operator"]
        if "allow_blank" in v:
            dv.allow_blank = v["allow_blank"]
        if "custom_error" in v:
            dv.error = v["custom_error"]
            dv.errorTitle = v["custom_title"]
        dv.showDropDown = False
        ws.add_data_validation(dv)
        dv.showErrorMessage = True
        dv.add(coords.full_range)


def add_validations(wb, ws, spec):
    for r in spec["open_ranges"]:
        coords = make_range(r)
        configure_validation(ws, coords, r)


def unlock_data_entry_cells(wb, ws, spec):
    for r in spec["open_ranges"]:
        # abs_start_col = r['title_cell'][0]
        # abs_start_row = int(r['title_cell'][1]) + 1
        coords = make_range(r)
        for rowndx in range(coords.range_start_row, MAX_ROWS):
            cell_reference = f"${coords.column}${rowndx}"
            cell = ws[cell_reference]
            cell.protection = Protection(locked=False)


def set_column_widths(wb, ws, spec):
    # Set he widths to something... sensible.
    # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    widths = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                widths[cell.column_letter] = max(
                    (widths.get(cell.column_letter, 0), len(str(cell.value)))
                )
    sum = 0
    for col, value in widths.items():
        sum += value
    avg = sum / len(widths)
    for col, value in widths.items():
        ws.column_dimensions[col].width = avg / 2
    for r in spec["open_ranges"]:
        column = r["title_cell"][0]
        if "width" in r:
            ws.column_dimensions[column].width = r["width"]


def generate_password():
    # https://pypi.org/project/xkcdpass/
    wordfile = xp.locate_wordfile()
    words = xp.generate_wordlist(wordfile=wordfile, min_length=7, max_length=9)
    correct_horse = "-".join(xp.generate_xkcdpassword(words).split(" "))
    return correct_horse


def set_wb_security(wb, password):
    # This is not doing what I want.
    # I cannot prevent the sheets from unlocking.
    wb.security = WorkbookProtection(workbookPassword=password, lockStructure=True)
    print(f"To unlock: {password}")


def save_workbook(wb, basename):
    wb.save(f"{basename}.xlsx")


def process_spec(spec):
    wb = create_empty_workbook()
    password = generate_password()
    for ndx, sheet in enumerate(spec["sheets"]):
        ws = create_protected_sheet(wb, sheet, password, ndx)
        if "cells_to_merge" in sheet:
            merge_adjacent_columns(ws, sheet["cells_to_merge"])
        process_open_ranges(wb, ws, sheet)
        add_validations(wb, ws, sheet)
        unlock_data_entry_cells(wb, ws, sheet)
        set_column_widths(wb, ws, sheet)
        process_single_cells(wb, ws, sheet)
        if "include_in_header" in sheet:
            apply_header_cell_style(ws, sheet["include_in_header"])

    set_wb_security(wb, password)
    return wb


if __name__ == "__main__":
    if len(sys.argv) == 2:
        maybe_filename = sys.argv[1]
        if maybe_filename.endswith("jsonnet"):
            spec = jsonnet_sheet_spec_to_json(maybe_filename)
            wb = process_spec(spec)
            save_workbook(wb, os.path.splitext(os.path.basename(sys.argv[1]))[0])
