from django.core.management.base import BaseCommand
from openpyxl import load_workbook
import os

class Command(BaseCommand):
    help = "Validates a merged workbook for basic structure and required fields"

    def handle(self, *args, **options):
        workbook_path = "/src/audit/test/MergedWorkbook_goodLMH.xlsx"

        if not os.path.exists(workbook_path):
            self.stderr.write(f"Workbook not found at: {workbook_path}")
            return

        wb = load_workbook(workbook_path, data_only=True)
        expected_sheets = [
            "general_information",
            "audit_information",
            "federal_awards",
            "findings_text",
            "findings_uniform_guidance",
            "corrective_action_plan",
            "notes_to_sefa",
        ]

        errors = []
        for sheet in expected_sheets:
            if sheet not in wb.sheetnames:
                errors.append(f"Missing expected sheet: {sheet}")
                continue

            ws = wb[sheet]
            if ws.max_row < 2:
                errors.append(f"Sheet '{sheet}' has too little data.")

        if errors:
            self.stdout.write("Validation Errors:")
            for err in errors:
                self.stdout.write(f" - {err}")
        else:
            self.stdout.write("Workbook passed basic validation checks.")
