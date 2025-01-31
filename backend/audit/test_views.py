import json
from datetime import datetime, timezone
from pathlib import Path
from tempfile import NamedTemporaryFile
from unittest.mock import patch

from audit.cross_validation.naming import SECTION_NAMES as SN
from audit.fixtures.excel import (
    ADDITIONAL_EINS_ENTRY_FIXTURES,
    ADDITIONAL_EINS_TEMPLATE,
    ADDITIONAL_UEIS_ENTRY_FIXTURES,
    ADDITIONAL_UEIS_TEMPLATE,
    CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES,
    CORRECTIVE_ACTION_PLAN_TEMPLATE,
    FEDERAL_AWARDS_ENTRY_FIXTURES,
    FEDERAL_AWARDS_TEMPLATE,
    FINDINGS_TEXT_ENTRY_FIXTURES,
    FINDINGS_TEXT_TEMPLATE,
    FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES,
    FINDINGS_UNIFORM_GUIDANCE_TEMPLATE,
    FORM_SECTIONS,
    NOTES_TO_SEFA_ENTRY_FIXTURES,
    NOTES_TO_SEFA_TEMPLATE,
    SECONDARY_AUDITORS_ENTRY_FIXTURES,
    SECONDARY_AUDITORS_TEMPLATE,
)
from audit.fixtures.single_audit_checklist import (
    fake_auditee_certification,
    fake_auditor_certification,
)
from audit.forms import AuditeeCertificationStep2Form, AuditorCertificationStep1Form
from audit.models import (
    Access,
    SingleAuditChecklist,
    SingleAuditReportFile,
    SubmissionEvent,
    generate_sac_report_id,
)
from audit.models.models import STATUS, ExcelFile
from audit.views import AuditeeCertificationStep2View, MySubmissions
from dissemination.models import FederalAward, General
from django.contrib.auth import get_user_model
from django.core.exceptions import PermissionDenied
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, RequestFactory, TestCase, TransactionTestCase
from django.urls import reverse
from faker import Faker
from model_bakery import baker
from openpyxl import load_workbook
from openpyxl.cell import Cell

User = get_user_model()

ACCESS_AND_SUBMISSION_PATH = reverse("report_submission:accessandsubmission")
AUDIT_JSON_FIXTURES = Path(__file__).parent / "fixtures" / "json"
EDIT_PATH = "audit:EditSubmission"
STATUSES = STATUS
SUBMISSIONS_PATH = reverse("audit:MySubmissions")

VALID_ELIGIBILITY_DATA = {
    "is_usa_based": True,
    "met_spending_threshold": True,
    "user_provided_organization_type": "state",
}

VALID_AUDITEE_INFO_DATA = {
    "auditee_uei": "ZQGGHJH74DW7",
    "auditee_fiscal_period_start": "2021-01-01",
    "auditee_fiscal_period_end": "2022-01-01",
    "auditee_name": "Tester",
}

VALID_ACCESS_AND_SUBMISSION_DATA = {
    "certifying_auditee_contact_fullname": "Fuller A. Namesmith",
    "certifying_auditee_contact_email": "a@a.com",
    "certifying_auditor_contact_fullname": "Fuller B. Namesmith",
    "certifying_auditor_contact_email": "b@b.com",
    "auditee_contacts_fullname": ["Fuller C. Namesmith"],
    "auditee_contacts_email": ["c@c.com"],
    "auditor_contacts_fullname": ["Fuller D. Namesmith"],
    "auditor_contacts_email": ["d@d.com"],
}


# Mocking the user login and file scan functions
def _mock_login_and_scan(client, mock_scan_file, **kwargs):
    """Helper function to mock the login and file scan functions"""
    user, sac = _make_user_and_sac(**kwargs)

    baker.make(Access, user=user, sac=sac)

    client.force_login(user)

    # mock the call to the external AV service
    mock_scan_file.return_value = MockHttpResponse(200, "clean!")

    return sac


def _client_post(client, view_str, kwargs=None, data=None):
    """Helper function for POST requests (does not force auth)"""
    kwargs, data = kwargs or {}, data or {}
    url = reverse(view_str, kwargs=kwargs)
    return client.post(url, data=data)


def _authed_post(client, user, view_str, kwargs=None, data=None):
    """Helper function for POST requests (forces auth)"""
    client.force_login(user=user)
    return _client_post(client, view_str, kwargs, data)


def _make_user_and_sac(**kwargs):
    """Helper function for to make a user and basic sac"""
    user = baker.make(User)
    sac = baker.make(SingleAuditChecklist, **kwargs)
    return user, sac


def _load_json(target):
    """Given a str or Path, load JSON from that target."""
    raw = Path(target).read_text(encoding="utf-8")
    return json.loads(raw)


def _mock_gen_report_id():
    """Helper function for generate a sac report id"""
    return generate_sac_report_id(end_date=datetime.now().date().isoformat())


def _merge_dict_seq(seq):
    """_merge_dict_seq([{1:2}, {3:4}]) => {1:2, 3:4}"""
    return {k: v for d in seq for k, v in d.items()}


def _build_auditor_cert_dict(certification: dict, signature: dict) -> dict:
    """Helper function for building a dictionary for auditor certification"""
    return {
        "auditor_certification": certification,
        "auditor_signature": signature,
    }


def _build_auditee_cert_dict(certification: dict, signature: dict) -> dict:
    """Helper function for building a dictionary for auditee certification"""
    return {
        "auditee_certification": certification,
        "auditee_signature": signature,
    }


def _just_uei(uei, fieldname):
    """
    _just_uei("whatever", "additional_eins") returns:
        {
            "AdditionalEINs": {"auditee_uei": "whatever"}
        }

    Given a UEI (a string) and a fieldname in
    audit.cross_validation.naming.SECTION_NAMES, return a structure with the camel-case
    version of that fieldname as the key and {"auditee_uei": uei} as the value.
    """
    return {SN[fieldname].camel_case: {"auditee_uei": uei}}


def _just_uei_workbooks(uei):
    """
    Given a UEI (a string), returns a dict containing all of the workbook snake_case
    field names with a value of {CamelCaseFieldName: {"auditee_uei": uei}} for each
    of those fields.
    """
    workbooks = {k: v for k, v in SN.items() if v.workbook_number}
    return {k: _just_uei(uei, k) for k in workbooks}


def _fake_audit_information():
    # TODO: consolidate all fixtures! This is a copy of a fixture from
    # intake_to_dissemination, which is not ideal.
    fake = Faker()

    return {
        "dollar_threshold": 10345.45,
        "gaap_results": json.dumps([fake.word()]),
        "is_going_concern_included": "Y" if fake.boolean() else "N",
        "is_internal_control_deficiency_disclosed": "Y" if fake.boolean() else "N",
        "is_internal_control_material_weakness_disclosed": (
            "Y" if fake.boolean() else "N"
        ),
        "is_material_noncompliance_disclosed": "Y" if fake.boolean() else "N",
        "is_aicpa_audit_guide_included": "Y" if fake.boolean() else "N",
        "is_low_risk_auditee": "Y" if fake.boolean() else "N",
        "agencies": json.dumps([fake.word()]),
    }


class RootPathTests(TestCase):
    """
    Do we return the correct response for both authenticated and unauthenticated
    users?
    """

    def test_unauthenticated(self):
        """Verify the root path returns 200, unauthenticated."""
        result = self.client.get("/")
        self.assertEqual(result.status_code, 200)

    def test_authenticated(self):
        """Verify the root path redirects, authenticated."""
        user = baker.make(User)
        self.client.force_login(user=user)
        result = self.client.get("/")
        self.assertEqual(result.status_code, 302)

    def test_no_robots(self):
        """Verify robots.txt returns 200"""
        result = self.client.get("/robots.txt")
        self.assertEqual(result.status_code, 200)


class MySubmissionsViewTests(TestCase):
    def setUp(self):
        """Setup function for users and client"""
        self.user = baker.make(User)
        self.user2 = baker.make(User)

        self.client = Client()

    def test_redirect_if_not_logged_in(self):
        """Test that accessing submission page redirects if user is not logged in"""
        result = self.client.get(SUBMISSIONS_PATH)
        self.assertAlmostEqual(result.status_code, 302)

    def test_no_submissions_returns_empty_list(self):
        """Test that an authenticated user with no submissions gets empty list"""
        self.client.force_login(user=self.user)
        data = MySubmissions.fetch_my_submissions(self.user)
        self.assertEqual(len(data), 0)

    def test_user_with_submissions_should_return_expected_data_columns(self):
        """Test that a user with submissions gets data with expected columns"""
        self.client.force_login(user=self.user)
        self.user.profile.entry_form_data = (
            VALID_ELIGIBILITY_DATA | VALID_AUDITEE_INFO_DATA
        )
        self.user.profile.save()
        self.client.post(
            ACCESS_AND_SUBMISSION_PATH, VALID_ACCESS_AND_SUBMISSION_DATA, format="json"
        )
        data = MySubmissions.fetch_my_submissions(self.user)
        self.assertGreater(len(data), 0)

        keys = data[0].keys()
        self.assertTrue("report_id" in keys)
        self.assertTrue("submission_status" in keys)
        self.assertTrue("auditee_uei" in keys)
        self.assertTrue("auditee_name" in keys)
        self.assertTrue("fiscal_year_end_date" in keys)

    def test_user_with_no_submissions_should_return_no_data(self):
        """Test that another user with no submissions gets no data"""
        self.client.force_login(user=self.user)
        self.user.profile.entry_form_data = (
            VALID_ELIGIBILITY_DATA | VALID_AUDITEE_INFO_DATA
        )
        self.user.profile.save()
        self.client.post(
            ACCESS_AND_SUBMISSION_PATH, VALID_ACCESS_AND_SUBMISSION_DATA, format="json"
        )
        data = MySubmissions.fetch_my_submissions(self.user2)
        self.assertEqual(len(data), 0)


class EditSubmissionViewTests(TestCase):
    def setUp(self):
        """Setup test factory, client, user, and report ID"""
        self.factory = RequestFactory()
        self.client = Client()
        self.user = baker.make(User)
        self.sac = baker.make(
            SingleAuditChecklist, submission_status=STATUS.READY_FOR_CERTIFICATION
        )
        self.url_name = "audit:EditSubmission"
        self.report_id = "TEST_REPORT_ID"
        self.url = reverse(
            "audit:EditSubmission", kwargs={"report_id": self.sac.report_id}
        )
        self.client.force_login(self.user)
        baker.make(
            "audit.Access",
            sac=self.sac,
            user=self.user,
            role="certifying_auditee_contact",
        )
        self.session = self.client.session

    def test_redirect_not_logged_in(self):
        """Test that accessing edit submission page redirects if not authenticated"""
        url = reverse(self.url_name, args=[self.report_id])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 302)

    def test_redirects_to_singleauditchecklist(self):
        """Test that accessing edit submission redirects to SAC view"""
        response = self.client.get(self.url)
        self.assertRedirects(
            response, reverse("singleauditchecklist", args=[self.sac.report_id])
        )


class SubmissionViewTests(TestCase):
    """
    Testing for the final step: submitting.
    """

    def setUp(self):
        """Set up test client, user, SAC, and URL"""
        self.client = Client()
        self.user = baker.make(User)
        self.sac = baker.make(
            SingleAuditChecklist, submission_status=STATUS.AUDITEE_CERTIFIED
        )
        self.url = reverse("audit:Submission", kwargs={"report_id": self.sac.report_id})
        self.client.force_login(self.user)
        baker.make(
            "audit.Access",
            sac=self.sac,
            user=self.user,
            role="certifying_auditee_contact",
        )

    def test_get_renders_template(self):
        """Test that GET renders the submission template with correct context"""
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "audit/submission.html")
        self.assertIn("report_id", response.context)
        self.assertIn("submission_status", response.context)
        self.assertEqual(response.context["report_id"], self.sac.report_id)
        self.assertEqual(
            response.context["submission_status"], self.sac.submission_status
        )

    def test_get_permission_denied_if_no_sac(self):
        """Test that GET returns 403 if SAC does not exist"""
        invalid_url = reverse("audit:Submission", kwargs={"report_id": "INVALID"})
        response = self.client.get(invalid_url)
        self.assertEqual(response.status_code, 403)

    def test_get_access_denied_for_unauthorized_user(self):
        """Test that GET returns 403 if user is unauthorized"""
        self.client.logout()
        response = self.client.get(self.url)
        self.assertTemplateUsed(response, "home.html")
        self.assertTrue(response.context["session_expired"])

    @patch("audit.models.SingleAuditChecklist.validate_full")
    @patch("audit.views.views.sac_transition")
    @patch("audit.views.views.remove_workbook_artifacts")
    @patch("audit.views.views.SingleAuditChecklist.disseminate")
    def test_post_successful(
        self, mock_disseminate, mock_remove, mock_transition, mock_validate
    ):
        """Test that a valid submission transitions SAC to a disseminated state"""
        mock_validate.return_value = []
        mock_disseminate.return_value = None
        response = self.client.post(self.url)

        mock_validate.assert_called_once()
        mock_disseminate.assert_called_once()
        mock_transition.assert_called_with(
            response.wsgi_request, self.sac, transition_to=STATUS.DISSEMINATED
        )
        mock_remove.assert_called_once()

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("audit:MySubmissions"))

    @patch("audit.views.views.SingleAuditChecklist.validate_full")
    @patch("audit.views.views.sac_transition")
    @patch("audit.views.views.SingleAuditChecklist.disseminate")
    def test_post_validation_errors(
        self, mock_disseminate, mock_transition, mock_validate
    ):
        """Test that validation errors are displayed if submission is invalid"""
        mock_validate.return_value = ["Error 1", "Error 2"]

        self.sac.submission_status = STATUS.AUDITEE_CERTIFIED
        self.sac.save()

        response = self.client.post(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "audit/cross-validation/cross-validation-results.html"
        )
        self.assertIn("errors", response.context)
        self.assertListEqual(response.context["errors"], ["Error 1", "Error 2"])

        mock_disseminate.assert_not_called()
        mock_transition.assert_not_called()

    @patch("audit.views.views.General.objects.get")
    @patch("audit.views.views.SingleAuditChecklist.validate_full")
    def test_post_transaction_error(self, mock_validate, mock_general_get):
        """Test that a transaction error during a submission is handled properly"""
        self.sac.submission_status = STATUS.AUDITEE_CERTIFIED
        self.sac.save()

        mock_validate.return_value = []
        mock_general_get.return_value = True

        response = self.client.post(self.url)

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("audit:MySubmissions"))

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("audit:MySubmissions"))

    def test_post_permission_denied_if_no_sac(self):
        """Test that POST returns 403 if SAC does not exist"""
        invalid_url = reverse("audit:Submission", kwargs={"report_id": "INVALID"})
        response = self.client.post(invalid_url)
        self.assertEqual(response.status_code, 403)

    def test_post_redirect(self):
        """
        The status should be "disseminated" after the post.
        The user should be redirected to the submissions table.
        """

        just_ueis = _just_uei_workbooks("TEST0001TEST")
        geninfofile = "general-information--test0001test--simple-pass.json"
        awardsfile = "federal-awards--test0001test--simple-pass.json"

        sac_data = just_ueis | {
            "auditee_certification": _build_auditee_cert_dict(
                *fake_auditee_certification()
            ),
            "auditor_certification": _build_auditor_cert_dict(
                *fake_auditor_certification()
            ),
            "audit_information": _fake_audit_information(),
            "federal_awards": _load_json(AUDIT_JSON_FIXTURES / awardsfile),
            "general_information": _load_json(AUDIT_JSON_FIXTURES / geninfofile),
            "submission_status": STATUSES.IN_PROGRESS,  # Temporarily required for SAR creation below
        }
        sac_data["notes_to_sefa"]["NotesToSefa"]["accounting_policies"] = "Exhaustive"
        sac_data["notes_to_sefa"]["NotesToSefa"]["is_minimis_rate_used"] = "Y"
        sac_data["notes_to_sefa"]["NotesToSefa"]["rate_explained"] = "At great length"
        sac_data["report_id"] = _mock_gen_report_id()
        user, sac = _make_user_and_sac(**sac_data)

        required_statuses = (
            STATUSES.AUDITEE_CERTIFIED,
            STATUSES.AUDITOR_CERTIFIED,
            STATUSES.READY_FOR_CERTIFICATION,
            STATUSES.CERTIFIED,
        )

        for rs in required_statuses:
            sac.transition_name.append(rs)
            sac.transition_date.append(datetime.now(timezone.utc))

        baker.make(SingleAuditReportFile, sac=sac)
        baker.make(Access, user=user, sac=sac, role="certifying_auditee_contact")
        sac.submission_status = STATUSES.AUDITEE_CERTIFIED
        sac.save()

        response = _authed_post(
            Client(),
            user,
            "audit:Submission",
            kwargs={"report_id": sac.report_id},
            data={},
        )
        sac_after = SingleAuditChecklist.objects.get(report_id=sac.report_id)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(sac_after.submission_status, STATUSES.DISSEMINATED)


class SubmissionGetTest(TestCase):
    def setUp(self):
        """Setup test user and client"""
        self.user = baker.make(User)
        self.client = Client()
        self.url = reverse("audit:MySubmissions")

    def testValidSubmission(self):
        """Test that a valid submission is displayed on the submissions page"""
        self.client.force_login(self.user)
        sac1 = baker.make(SingleAuditChecklist, submission_status=STATUS.IN_PROGRESS)
        sac2 = baker.make(SingleAuditChecklist, submission_status=STATUS.DISSEMINATED)
        baker.make(Access, user=self.user, sac=sac1)
        baker.make(Access, user=self.user, sac=sac2)

        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "audit/audit_submissions/audit_submissions.html"
        )


class SubmissionStatusTests(TransactionTestCase):
    """
    Tests the expected order of progression for submission_status.

    Does not yet test that only users with the appropriate permissions can
    access the relevant pages.

    Does not test attempts at incorrect progressions, as these are handled by
    tests in test_models.py.
    """

    def setUp(self):
        """Setup user and client"""
        self.user = baker.make(User)
        self.client = Client()

    def test_ready_for_certification(self):
        """
        Test that the submitting user can mark the submission as ready for certification
        """
        self.client.force_login(user=self.user)
        self.user.profile.entry_form_data = (
            VALID_ELIGIBILITY_DATA | VALID_AUDITEE_INFO_DATA
        )
        self.user.profile.save()
        self.client.post(
            ACCESS_AND_SUBMISSION_PATH, VALID_ACCESS_AND_SUBMISSION_DATA, format="json"
        )
        data = MySubmissions.fetch_my_submissions(self.user)
        self.assertGreater(len(data), 0)
        self.assertEqual(data[0]["submission_status"], STATUSES.IN_PROGRESS)
        report_id = data[0]["report_id"]

        # Update the SAC so that it will pass overall validation:
        geninfofile = "general-information--test0001test--simple-pass.json"
        awardsfile = "federal-awards--test0001test--simple-pass.json"
        just_ueis = _just_uei_workbooks("TEST0001TEST")
        sac_data = just_ueis | {
            "auditee_certification": _build_auditee_cert_dict(
                *fake_auditee_certification()
            ),
            "auditor_certification": _build_auditor_cert_dict(
                *fake_auditor_certification()
            ),
            "audit_information": _fake_audit_information(),
            "federal_awards": _load_json(AUDIT_JSON_FIXTURES / awardsfile),
            "general_information": _load_json(AUDIT_JSON_FIXTURES / geninfofile),
        }
        sac_data["notes_to_sefa"]["NotesToSefa"]["accounting_policies"] = "Exhaustive"
        sac_data["notes_to_sefa"]["NotesToSefa"]["is_minimis_rate_used"] = "Y"
        sac_data["notes_to_sefa"]["NotesToSefa"]["rate_explained"] = "At great length"

        sac = SingleAuditChecklist.objects.get(report_id=report_id)
        for field, value in sac_data.items():
            setattr(sac, field, value)
        baker.make(SingleAuditReportFile, sac=sac)
        sac.save()

        self.client.post(f"/audit/ready-for-certification/{report_id}", data={})
        data = MySubmissions.fetch_my_submissions(self.user)

        self.assertEqual(data[0]["submission_status"], STATUSES.READY_FOR_CERTIFICATION)

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be LOCKED_FOR_CERTIFICATION
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.LOCKED_FOR_CERTIFICATION,
        )

    def test_unlock_after_certification(self):
        """
        Test that any user can move the submission back to in progress.
        """
        self.client.force_login(user=self.user)
        self.user.profile.entry_form_data = (
            VALID_ELIGIBILITY_DATA | VALID_AUDITEE_INFO_DATA
        )
        self.user.profile.save()
        self.client.post(
            ACCESS_AND_SUBMISSION_PATH, VALID_ACCESS_AND_SUBMISSION_DATA, format="json"
        )
        data = MySubmissions.fetch_my_submissions(self.user)
        self.assertGreater(len(data), 0)
        self.assertEqual(data[0]["submission_status"], STATUSES.IN_PROGRESS)
        report_id = data[0]["report_id"]

        # Update the SAC so that it will pass overall validation:
        geninfofile = "general-information--test0001test--simple-pass.json"
        awardsfile = "federal-awards--test0001test--simple-pass.json"
        just_ueis = _just_uei_workbooks("TEST0001TEST")
        sac_data = just_ueis | {
            "auditee_certification": _build_auditee_cert_dict(
                *fake_auditee_certification()
            ),
            "auditor_certification": _build_auditor_cert_dict(
                *fake_auditor_certification()
            ),
            "audit_information": _fake_audit_information(),
            "federal_awards": _load_json(AUDIT_JSON_FIXTURES / awardsfile),
            "general_information": _load_json(AUDIT_JSON_FIXTURES / geninfofile),
        }
        sac_data["notes_to_sefa"]["NotesToSefa"]["accounting_policies"] = "Exhaustive"
        sac_data["notes_to_sefa"]["NotesToSefa"]["is_minimis_rate_used"] = "Y"
        sac_data["notes_to_sefa"]["NotesToSefa"]["rate_explained"] = "At great length"

        sac = SingleAuditChecklist.objects.get(report_id=report_id)
        for field, value in sac_data.items():
            setattr(sac, field, value)
        baker.make(SingleAuditReportFile, sac=sac)
        sac.save()

        self.client.post(f"/audit/ready-for-certification/{report_id}", data={})
        data = MySubmissions.fetch_my_submissions(self.user)

        self.assertEqual(data[0]["submission_status"], STATUSES.READY_FOR_CERTIFICATION)

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be LOCKED_FOR_CERTIFICATION
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.LOCKED_FOR_CERTIFICATION,
        )

        postdata = {"unlock_after_certification": True}
        self.client.post(
            f"/audit/unlock-after-certification/{report_id}", data=postdata
        )
        data = MySubmissions.fetch_my_submissions(self.user)

        self.assertEqual(data[0]["submission_status"], STATUSES.IN_PROGRESS)

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be UNLOCKED_AFTER_CERTIFICATION
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.UNLOCKED_AFTER_CERTIFICATION,
        )

    def test_auditor_certification(self):
        """
        Test that certifying auditor contacts can provide auditor certification
        """
        data_step_1, data_step_2 = fake_auditor_certification()
        just_ueis = _just_uei_workbooks("TEST0001TEST")
        sac_data = just_ueis | {"submission_status": STATUSES.READY_FOR_CERTIFICATION}

        user, sac = _make_user_and_sac(**sac_data)
        baker.make(Access, sac=sac, user=user, role="certifying_auditor_contact")

        kwargs = {"report_id": sac.report_id}
        _authed_post(
            self.client,
            user,
            "audit:AuditorCertification",
            kwargs=kwargs,
            data=data_step_1,
        )
        _authed_post(
            self.client,
            user,
            "audit:AuditorCertificationConfirm",
            kwargs=kwargs,
            data=data_step_1 | data_step_2,
        )

        updated_sac = SingleAuditChecklist.objects.get(report_id=sac.report_id)

        self.assertEqual(updated_sac.submission_status, STATUSES.AUDITOR_CERTIFIED)

        submission_events = SubmissionEvent.objects.filter(sac=sac)
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.AUDITOR_CERTIFICATION_COMPLETED,
        )

    def test_auditee_certification(self):
        """
        Test that certifying auditee contacts can provide auditee certification
        """
        data_step_1, data_step_2 = fake_auditee_certification()
        just_ueis = _just_uei_workbooks("TEST0001TEST")
        sac_data = just_ueis | {"submission_status": STATUSES.AUDITOR_CERTIFIED}
        user, sac = _make_user_and_sac(**sac_data)
        baker.make(Access, sac=sac, user=user, role="certifying_auditee_contact")

        kwargs = {"report_id": sac.report_id}
        _authed_post(
            self.client,
            user,
            "audit:AuditeeCertification",
            kwargs=kwargs,
            data=data_step_1,
        )
        _authed_post(
            self.client,
            user,
            "audit:AuditeeCertificationConfirm",
            kwargs=kwargs,
            data=data_step_1 | data_step_2,
        )

        updated_sac = SingleAuditChecklist.objects.get(report_id=sac.report_id)

        self.assertEqual(updated_sac.submission_status, STATUSES.AUDITEE_CERTIFIED)

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be AUDITEE_CERTIFICATION_COMPLETED
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.AUDITEE_CERTIFICATION_COMPLETED,
        )

    def test_submission(self):
        """
        Test that certifying auditee contacts can perform submission
        """
        just_ueis = _just_uei_workbooks("TEST0001TEST")
        geninfofile = "general-information--test0001test--simple-pass.json"
        awardsfile = "federal-awards--test0001test--simple-pass.json"

        sac_data = just_ueis | {
            "auditee_certification": _build_auditee_cert_dict(
                *fake_auditee_certification()
            ),
            "auditor_certification": _build_auditor_cert_dict(
                *fake_auditor_certification()
            ),
            "audit_information": _fake_audit_information(),
            "federal_awards": _load_json(AUDIT_JSON_FIXTURES / awardsfile),
            "general_information": _load_json(AUDIT_JSON_FIXTURES / geninfofile),
            "submission_status": STATUSES.IN_PROGRESS,  # Temporarily required for SAR creation below
        }
        sac_data["notes_to_sefa"]["NotesToSefa"]["accounting_policies"] = "Exhaustive"
        sac_data["notes_to_sefa"]["NotesToSefa"]["is_minimis_rate_used"] = "Y"
        sac_data["notes_to_sefa"]["NotesToSefa"]["rate_explained"] = "At great length"
        sac_data["report_id"] = _mock_gen_report_id()
        user, sac = _make_user_and_sac(**sac_data)

        required_statuses = (
            STATUSES.AUDITEE_CERTIFIED,
            STATUSES.AUDITOR_CERTIFIED,
            STATUSES.READY_FOR_CERTIFICATION,
            STATUSES.CERTIFIED,
        )

        for rs in required_statuses:
            sac.transition_name.append(rs)
            sac.transition_date.append(datetime.now(timezone.utc))

        baker.make(SingleAuditReportFile, sac=sac)
        baker.make(Access, sac=sac, user=user, role="certifying_auditee_contact")
        sac.submission_status = STATUSES.AUDITEE_CERTIFIED
        sac.save()

        kwargs = {"report_id": sac.report_id}
        _authed_post(self.client, user, "audit:Submission", kwargs=kwargs)

        updated_sac = SingleAuditChecklist.objects.get(report_id=sac.report_id)

        self.assertEqual(updated_sac.submission_status, STATUSES.DISSEMINATED)

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be SUBMITTED
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.DISSEMINATED,
        )

    def test_submission_race_condition(self):
        """
        Test that certifying auditee contacts can perform submission
        """
        just_ueis = _just_uei_workbooks("TEST0001TEST")
        geninfofile = "general-information--test0001test--simple-pass.json"
        awardsfile = "federal-awards--test0001test--simple-pass.json"

        sac_data = just_ueis | {
            "auditee_certification": _build_auditee_cert_dict(
                *fake_auditee_certification()
            ),
            "auditor_certification": _build_auditor_cert_dict(
                *fake_auditor_certification()
            ),
            "audit_information": _fake_audit_information(),
            "federal_awards": _load_json(AUDIT_JSON_FIXTURES / awardsfile),
            "general_information": _load_json(AUDIT_JSON_FIXTURES / geninfofile),
            "submission_status": STATUSES.AUDITEE_CERTIFIED,
        }
        sac_data = just_ueis | {
            "auditee_certification": _build_auditee_cert_dict(
                *fake_auditee_certification()
            ),
            "auditor_certification": _build_auditor_cert_dict(
                *fake_auditor_certification()
            ),
            "audit_information": _fake_audit_information(),
            "federal_awards": _load_json(AUDIT_JSON_FIXTURES / awardsfile),
            "general_information": _load_json(AUDIT_JSON_FIXTURES / geninfofile),
            "submission_status": STATUSES.AUDITEE_CERTIFIED,
        }
        sac_data["notes_to_sefa"]["NotesToSefa"]["accounting_policies"] = "Exhaustive"
        sac_data["notes_to_sefa"]["NotesToSefa"]["is_minimis_rate_used"] = "Y"
        sac_data["notes_to_sefa"]["NotesToSefa"]["rate_explained"] = "At great length"
        user, sac = _make_user_and_sac(**sac_data)

        required_statuses = (
            STATUSES.AUDITEE_CERTIFIED,
            STATUSES.AUDITOR_CERTIFIED,
            STATUSES.READY_FOR_CERTIFICATION,
            STATUSES.CERTIFIED,
        )

        for rs in required_statuses:
            sac.transition_name.append(rs)
            sac.transition_date.append(datetime.now(timezone.utc))

        sac.save()

        baker.make(Access, sac=sac, user=user, role="certifying_auditee_contact")

        # For this test, insert a matching report_id into general so that attempts to
        # disseminate SACs with this report_id will fail:
        general = baker.make(General, report_id=sac.report_id)
        general.save()

        kwargs = {"report_id": sac.report_id}
        _authed_post(self.client, user, "audit:Submission", kwargs=kwargs)

        # The above post should fail on dissemination and put nothing in
        # dissemination.FederalAward, which, since this is a test context, should
        # therefore be empty.
        self.assertEqual(0, FederalAward.objects.count())


class MockHttpResponse:
    def __init__(self, status_code, text):
        self.status_code = status_code
        self.text = text


def _set_by_name(workbook, name, value, row_offset=0):
    definition = workbook.defined_names[name]

    sheet_title, cell_coord = next(definition.destinations)

    sheet = workbook[sheet_title]
    cell_range = sheet[cell_coord]

    if isinstance(cell_range, Cell):
        cell_range.value = value
    else:
        cell_range[row_offset][0].value = value


def _add_entry(workbook, row_offset, entry):
    for key, value in entry.items():
        _set_by_name(workbook, key, value, row_offset)


class ExcelFileHandlerViewTests(TestCase):
    GOOD_UEI = "AAA123456BBB"
    BAD_UEI = "THIS_IS_NOT_A_VALID_UEI"

    def test_login_required(self):
        """When an unauthenticated request is made"""

        for form_section in FORM_SECTIONS:
            response = self.client.post(
                reverse(
                    f"audit:{form_section}",
                    kwargs={"report_id": "12345", "form_section": form_section},
                )
            )

            self.assertTemplateUsed(response, "home.html")
            self.assertTrue(response.context["session_expired"])

    def test_bad_report_id_returns_403(self):
        """When a request is made for a malformed or nonexistent report_id, a 403 error should be returned"""
        user = baker.make(User)

        self.client.force_login(user)

        for form_section in FORM_SECTIONS:
            response = self.client.post(
                reverse(
                    f"audit:{form_section}",
                    kwargs={
                        "report_id": "this is not a report id",
                        "form_section": form_section,
                    },
                )
            )

            self.assertEqual(response.status_code, 403)

    def test_inaccessible_audit_returns_403(self):
        """When a request is made for an audit that is inaccessible for this user, a 403 error should be returned"""
        user, sac = _make_user_and_sac()

        self.client.force_login(user)
        for form_section in FORM_SECTIONS:
            response = self.client.post(
                reverse(
                    f"audit:{form_section}",
                    kwargs={"report_id": sac.report_id, "form_section": form_section},
                )
            )

            self.assertEqual(response.status_code, 403)

    def test_no_file_attached_returns_400(self):
        """When a request is made with no file attached, a 400 error should be returned"""
        user, sac = _make_user_and_sac()
        baker.make(Access, user=user, sac=sac)

        self.client.force_login(user)

        for form_section in FORM_SECTIONS:
            response = self.client.post(
                reverse(
                    f"audit:{form_section}",
                    kwargs={"report_id": sac.report_id, "form_section": form_section},
                )
            )

            self.assertEqual(response.status_code, 400)

    def test_invalid_file_upload_returns_400(self):
        """When an invalid Excel file is uploaded, a 400 error should be returned"""
        user, sac = _make_user_and_sac()
        baker.make(Access, user=user, sac=sac)

        self.client.force_login(user)

        file = SimpleUploadedFile("not-excel.txt", b"asdf", "text/plain")

        for form_section in FORM_SECTIONS:
            response = self.client.post(
                reverse(
                    f"audit:{form_section}",
                    kwargs={"report_id": sac.report_id, "form_section": form_section},
                ),
                data={"FILES": file},
            )

            self.assertEqual(response.status_code, 400)

    # @patch("audit.validators._scan_file")
    # def test_valid_file_upload_for_federal_awards(self, mock_scan_file):
    #     """When a valid Excel file is uploaded, the file should be stored and the SingleAuditChecklist should be updated to include the uploaded Federal Awards data"""

    #     sac = _mock_login_and_scan(
    #         self.client,
    #         mock_scan_file,
    #         report_id=_mock_gen_report_id(),
    #     )
    #     test_data = json.loads(
    #         FEDERAL_AWARDS_ENTRY_FIXTURES.read_text(encoding="utf-8")
    #     )

    #     # add valid data to the workbook
    #     workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)
    #     _set_by_name(workbook, "total_amount_expended", test_data[0]["amount_expended"])
    #     _set_by_name(workbook, "auditee_uei", ExcelFileHandlerViewTests.GOOD_UEI)
    #     _set_by_name(workbook, "section_name", FORM_SECTIONS.FEDERAL_AWARDS)
    #     _add_entry(workbook, 0, test_data[0])

    #     with NamedTemporaryFile(suffix=".xlsx") as tmp:
    #         workbook.save(tmp.name)
    #         tmp.seek(0)

    #         with open(tmp.name, "rb") as excel_file:
    #             response = self.client.post(
    #                 reverse(
    #                     f"audit:{FORM_SECTIONS.FEDERAL_AWARDS}",
    #                     kwargs={
    #                         "report_id": sac.report_id,
    #                         "form_section": FORM_SECTIONS.FEDERAL_AWARDS,
    #                     },
    #                 ),
    #                 data={"FILES": excel_file},
    #             )

    #             self.assertEqual(response.status_code, 302)

    #             updated_sac = SingleAuditChecklist.objects.get(pk=sac.id)

    #             self.assertEqual(
    #                 updated_sac.federal_awards["FederalAwards"]["auditee_uei"],
    #                 ExcelFileHandlerViewTests.GOOD_UEI,
    #             )
    #             self.assertEqual(
    #                 len(updated_sac.federal_awards["FederalAwards"]["federal_awards"]),
    #                 1,
    #             )

    #             federal_awards_entry = updated_sac.federal_awards["FederalAwards"][
    #                 "federal_awards"
    #             ][0]

    #             self.assertEqual(
    #                 federal_awards_entry["cluster"]["cluster_name"],
    #                 test_data[0]["cluster_name"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["direct_or_indirect_award"]["is_direct"],
    #                 test_data[0]["is_direct"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["program"]["is_major"],
    #                 test_data[0]["is_major"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["program"]["federal_agency_prefix"],
    #                 test_data[0]["federal_agency_prefix"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["program"]["three_digit_extension"],
    #                 test_data[0]["three_digit_extension"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["program"]["amount_expended"],
    #                 test_data[0]["amount_expended"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["program"]["program_name"],
    #                 test_data[0]["program_name"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["loan_or_loan_guarantee"]["is_guaranteed"],
    #                 test_data[0]["is_guaranteed"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["program"]["number_of_audit_findings"],
    #                 test_data[0]["number_of_audit_findings"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["program"]["audit_report_type"],
    #                 test_data[0]["audit_report_type"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["subrecipients"]["is_passed"],
    #                 test_data[0]["is_passed"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["subrecipients"]["subrecipient_amount"],
    #                 test_data[0]["subrecipient_amount"],
    #             )
    #             self.assertEqual(
    #                 federal_awards_entry["direct_or_indirect_award"]["entities"],
    #                 [
    #                     {
    #                         "passthrough_name": "A",
    #                         "passthrough_identifying_number": "1",
    #                     },
    #                     {
    #                         "passthrough_name": "B",
    #                         "passthrough_identifying_number": "2",
    #                     },
    #                 ],
    #             )

    #     submission_events = SubmissionEvent.objects.filter(sac=sac)

    #     # the most recent event should be FEDERAL_AWARDS_UPDATED
    #     event_count = len(submission_events)
    #     self.assertGreaterEqual(event_count, 1)
    #     self.assertEqual(
    #         submission_events[event_count - 1].event,
    #         SubmissionEvent.EventType.FEDERAL_AWARDS_UPDATED,
    #     )

    @patch("audit.validators._scan_file")
    def test_valid_file_upload_for_corrective_action_plan(self, mock_scan_file):
        """When a valid Excel file is uploaded, the file should be stored and the SingleAuditChecklist should be updated to include the uploaded Corrective Action Plan data"""

        test_uei = "AAA12345678X"
        sac = _mock_login_and_scan(
            self.client,
            mock_scan_file,
            report_id=_mock_gen_report_id(),
        )
        test_data = json.loads(
            CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES.read_text(encoding="utf-8")
        )

        # add valid data to the workbook
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)
        _set_by_name(workbook, "auditee_uei", test_uei)
        _set_by_name(workbook, "section_name", FORM_SECTIONS.CORRECTIVE_ACTION_PLAN)

        _add_entry(workbook, 0, test_data[0])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)

            with open(tmp.name, "rb") as excel_file:
                response = self.client.post(
                    reverse(
                        f"audit:{FORM_SECTIONS.CORRECTIVE_ACTION_PLAN}",
                        kwargs={
                            "report_id": sac.report_id,
                            "form_section": FORM_SECTIONS.CORRECTIVE_ACTION_PLAN,
                        },
                    ),
                    data={"FILES": excel_file},
                )

                self.assertEqual(response.status_code, 302)

                updated_sac = SingleAuditChecklist.objects.get(pk=sac.id)

                self.assertEqual(
                    updated_sac.corrective_action_plan["CorrectiveActionPlan"][
                        "auditee_uei"
                    ],
                    test_uei,
                )

                self.assertEqual(
                    len(
                        updated_sac.corrective_action_plan["CorrectiveActionPlan"][
                            "corrective_action_plan_entries"
                        ]
                    ),
                    1,
                )

                corrective_action_plan_entry = updated_sac.corrective_action_plan[
                    "CorrectiveActionPlan"
                ]["corrective_action_plan_entries"][0]

                self.assertEqual(
                    corrective_action_plan_entry["planned_action"],
                    test_data[0]["planned_action"],
                )
                self.assertEqual(
                    corrective_action_plan_entry["reference_number"],
                    test_data[0]["reference_number"],
                )

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be CORRECTIVE_ACTION_PLAN_UPDATED
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.CORRECTIVE_ACTION_PLAN_UPDATED,
        )

    @patch("audit.validators._scan_file")
    def test_valid_file_upload_for_findings_uniform_guidance(self, mock_scan_file):
        """When a valid Excel file is uploaded, the file should be stored and the SingleAuditChecklist should be updated to include the uploaded Findings Uniform Guidance data"""

        sac = _mock_login_and_scan(
            self.client,
            mock_scan_file,
            report_id=_mock_gen_report_id(),
        )
        test_data = json.loads(
            FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES.read_text(encoding="utf-8")
        )

        # add valid data to the workbook
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)
        _set_by_name(workbook, "auditee_uei", ExcelFileHandlerViewTests.GOOD_UEI)
        _set_by_name(workbook, "section_name", FORM_SECTIONS.FINDINGS_UNIFORM_GUIDANCE)
        _add_entry(workbook, 0, test_data[0])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)

            with open(tmp.name, "rb") as excel_file:
                response = self.client.post(
                    reverse(
                        f"audit:{FORM_SECTIONS.FINDINGS_UNIFORM_GUIDANCE}",
                        kwargs={
                            "report_id": sac.report_id,
                            "form_section": FORM_SECTIONS.FINDINGS_UNIFORM_GUIDANCE,
                        },
                    ),
                    data={"FILES": excel_file},
                )

                self.assertEqual(response.status_code, 302)

                updated_sac = SingleAuditChecklist.objects.get(pk=sac.id)

                self.assertEqual(
                    updated_sac.findings_uniform_guidance["FindingsUniformGuidance"][
                        "auditee_uei"
                    ],
                    ExcelFileHandlerViewTests.GOOD_UEI,
                )

                self.assertEqual(
                    len(
                        updated_sac.findings_uniform_guidance[
                            "FindingsUniformGuidance"
                        ]["findings_uniform_guidance_entries"]
                    ),
                    1,
                )

                findings_entries = updated_sac.findings_uniform_guidance[
                    "FindingsUniformGuidance"
                ]["findings_uniform_guidance_entries"][0]

                self.assertEqual(
                    findings_entries["program"]["award_reference"],
                    test_data[0]["award_reference"],
                )
                self.assertEqual(
                    findings_entries["program"]["compliance_requirement"],
                    test_data[0]["compliance_requirement"],
                )
                self.assertEqual(
                    findings_entries["findings"]["repeat_prior_reference"],
                    test_data[0]["repeat_prior_reference"],
                )
                self.assertEqual(
                    findings_entries["findings"]["reference_number"],
                    test_data[0]["reference_number"],
                )
                self.assertEqual(
                    findings_entries["modified_opinion"],
                    test_data[0]["modified_opinion"],
                )

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be FINDINGS_UNIFORM_GUIDANCE
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.FINDINGS_UNIFORM_GUIDANCE_UPDATED,
        )

    @patch("audit.validators._scan_file")
    def test_valid_file_upload_for_findings_text(self, mock_scan_file):
        """When a valid Excel file is uploaded, the file should be stored and the SingleAuditChecklist should be updated to include the uploaded Findings Text data"""

        sac = _mock_login_and_scan(
            self.client,
            mock_scan_file,
            report_id=_mock_gen_report_id(),
        )
        test_data = json.loads(FINDINGS_TEXT_ENTRY_FIXTURES.read_text(encoding="utf-8"))

        # add valid data to the workbook
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)
        _set_by_name(workbook, "auditee_uei", ExcelFileHandlerViewTests.GOOD_UEI)
        _set_by_name(workbook, "section_name", FORM_SECTIONS.FINDINGS_TEXT)
        _add_entry(workbook, 0, test_data[0])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)

            with open(tmp.name, "rb") as excel_file:
                response = self.client.post(
                    reverse(
                        f"audit:{FORM_SECTIONS.FINDINGS_TEXT}",
                        kwargs={
                            "report_id": sac.report_id,
                            "form_section": FORM_SECTIONS.FINDINGS_TEXT,
                        },
                    ),
                    data={"FILES": excel_file},
                )

                self.assertEqual(response.status_code, 302)

                updated_sac = SingleAuditChecklist.objects.get(pk=sac.id)

                self.assertEqual(
                    updated_sac.findings_text["FindingsText"]["auditee_uei"],
                    ExcelFileHandlerViewTests.GOOD_UEI,
                )

                self.assertEqual(
                    len(
                        updated_sac.findings_text["FindingsText"][
                            "findings_text_entries"
                        ]
                    ),
                    1,
                )

                findings_entries = updated_sac.findings_text["FindingsText"][
                    "findings_text_entries"
                ][0]

                self.assertEqual(
                    findings_entries["contains_chart_or_table"],
                    test_data[0]["contains_chart_or_table"],
                )
                self.assertEqual(
                    findings_entries["text_of_finding"],
                    test_data[0]["text_of_finding"],
                )
                self.assertEqual(
                    findings_entries["reference_number"],
                    test_data[0]["reference_number"],
                )

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be FEDERAL_AWARDS_AUDIT_FINDINGS_TEXT_UPDATED
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.FEDERAL_AWARDS_AUDIT_FINDINGS_TEXT_UPDATED,
        )

    @patch("audit.validators._scan_file")
    def test_valid_file_upload_for_secondary_auditors(self, mock_scan_file):
        """When a valid Excel file is uploaded, the file should be stored and the SingleAuditChecklist should be updated to include the uploaded secondary auditors data"""

        sac = _mock_login_and_scan(
            self.client,
            mock_scan_file,
            report_id=_mock_gen_report_id(),
        )
        test_data = json.loads(
            SECONDARY_AUDITORS_ENTRY_FIXTURES.read_text(encoding="utf-8")
        )

        # add valid data to the workbook
        workbook = load_workbook(SECONDARY_AUDITORS_TEMPLATE, data_only=True)
        _set_by_name(workbook, "auditee_uei", ExcelFileHandlerViewTests.GOOD_UEI)
        _set_by_name(workbook, "section_name", FORM_SECTIONS.SECONDARY_AUDITORS)
        _add_entry(workbook, 0, test_data[0])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)

            with open(tmp.name, "rb") as excel_file:
                response = self.client.post(
                    reverse(
                        f"audit:{FORM_SECTIONS.SECONDARY_AUDITORS}",
                        kwargs={
                            "report_id": sac.report_id,
                            "form_section": FORM_SECTIONS.SECONDARY_AUDITORS,
                        },
                    ),
                    data={"FILES": excel_file},
                )

                self.assertEqual(response.status_code, 302)

                updated_sac = SingleAuditChecklist.objects.get(pk=sac.id)

                self.assertEqual(
                    updated_sac.secondary_auditors["SecondaryAuditors"]["auditee_uei"],
                    ExcelFileHandlerViewTests.GOOD_UEI,
                )

                self.assertEqual(
                    len(
                        updated_sac.secondary_auditors["SecondaryAuditors"][
                            "secondary_auditors_entries"
                        ]
                    ),
                    1,
                )

                secondary_auditors_entries = updated_sac.secondary_auditors[
                    "SecondaryAuditors"
                ]["secondary_auditors_entries"][0]

                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_name"],
                    test_data[0]["secondary_auditor_name"],
                )
                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_ein"],
                    test_data[0]["secondary_auditor_ein"],
                )
                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_address_street"],
                    test_data[0]["secondary_auditor_address_street"],
                )
                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_address_city"],
                    test_data[0]["secondary_auditor_address_city"],
                )
                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_address_state"],
                    test_data[0]["secondary_auditor_address_state"],
                )
                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_address_zipcode"],
                    test_data[0]["secondary_auditor_address_zipcode"],
                )
                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_contact_name"],
                    test_data[0]["secondary_auditor_contact_name"],
                )
                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_contact_title"],
                    test_data[0]["secondary_auditor_contact_title"],
                )
                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_contact_phone"],
                    test_data[0]["secondary_auditor_contact_phone"],
                )
                self.assertEqual(
                    secondary_auditors_entries["secondary_auditor_contact_email"],
                    test_data[0]["secondary_auditor_contact_email"],
                )

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be SECONDARY_AUDITORS_UPDATED
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.SECONDARY_AUDITORS_UPDATED,
        )

    @patch("audit.validators._scan_file")
    def test_late_file_upload(self, mock_scan_file):
        """When a valid Excel file is uploaded after the submission has been locked, the upload should be rejected"""

        test_cases = [
            (
                FEDERAL_AWARDS_ENTRY_FIXTURES,
                FEDERAL_AWARDS_TEMPLATE,
                FORM_SECTIONS.FEDERAL_AWARDS,
            ),
            (
                CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES,
                CORRECTIVE_ACTION_PLAN_TEMPLATE,
                FORM_SECTIONS.CORRECTIVE_ACTION_PLAN,
            ),
            (
                FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES,
                FINDINGS_UNIFORM_GUIDANCE_TEMPLATE,
                FORM_SECTIONS.FINDINGS_UNIFORM_GUIDANCE,
            ),
            (
                FINDINGS_TEXT_ENTRY_FIXTURES,
                FINDINGS_TEXT_TEMPLATE,
                FORM_SECTIONS.FINDINGS_TEXT,
            ),
            (
                SECONDARY_AUDITORS_ENTRY_FIXTURES,
                SECONDARY_AUDITORS_TEMPLATE,
                FORM_SECTIONS.SECONDARY_AUDITORS,
            ),
        ]

        for test_case in test_cases:
            with self.subTest():
                fixtures, template, section = test_case

                sac = _mock_login_and_scan(
                    self.client,
                    mock_scan_file,
                    report_id=_mock_gen_report_id(),
                    submission_status=STATUSES.READY_FOR_CERTIFICATION,
                )

                test_data = json.loads(fixtures.read_text(encoding="utf-8"))

                # add valid data to the workbook
                workbook = load_workbook(template, data_only=True)
                _set_by_name(
                    workbook, "auditee_uei", ExcelFileHandlerViewTests.GOOD_UEI
                )
                _set_by_name(workbook, "section_name", section)
                _add_entry(workbook, 0, test_data[0])

                with NamedTemporaryFile(suffix=".xlsx") as tmp:
                    workbook.save(tmp.name)
                    tmp.seek(0)

                    with open(tmp.name, "rb") as excel_file:
                        response = self.client.post(
                            reverse(
                                f"audit:{section}",
                                kwargs={
                                    "report_id": sac.report_id,
                                    "form_section": section,
                                },
                            ),
                            data={"FILES": excel_file},
                        )

                        self.assertEqual(response.status_code, 400)
                        self.assertIn(
                            "no_late_changes", response.content.decode("utf-8")
                        )

    def test_get_login_required(self):
        """Test that uploading files requires user authentication"""
        for form_section in FORM_SECTIONS:
            response = self.client.get(
                reverse(
                    f"audit:{form_section}",
                    kwargs={"report_id": "12345", "form_section": form_section},
                )
            )
            self.assertTemplateUsed(response, "home.html")
            self.assertTrue(response.context["session_expired"])

    def test_get_bad_report_id_returns_403(self):
        """Test that uploading with a malformed or nonexistant report_id reutrns 403"""
        user = baker.make(User)
        self.client.force_login(user)

        for form_section in FORM_SECTIONS:
            response = self.client.get(
                reverse(
                    f"audit:{form_section}",
                    kwargs={
                        "report_id": "this is not a report id",
                        "form_section": form_section,
                    },
                )
            )
            self.assertEqual(response.status_code, 403)

    def test_excel_file_not_saved_on_validation_failure(self):
        """Ensure that the Excel file is not saved to the database if validation fails."""

        user, sac = _make_user_and_sac()
        baker.make(Access, user=user, sac=sac)

        self.client.force_login(user)

        test_cases = [
            (
                FEDERAL_AWARDS_ENTRY_FIXTURES,
                FEDERAL_AWARDS_TEMPLATE,
                FORM_SECTIONS.FEDERAL_AWARDS,
            ),
            (
                CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES,
                CORRECTIVE_ACTION_PLAN_TEMPLATE,
                FORM_SECTIONS.CORRECTIVE_ACTION_PLAN,
            ),
            (
                FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES,
                FINDINGS_UNIFORM_GUIDANCE_TEMPLATE,
                FORM_SECTIONS.FINDINGS_UNIFORM_GUIDANCE,
            ),
            (
                FINDINGS_TEXT_ENTRY_FIXTURES,
                FINDINGS_TEXT_TEMPLATE,
                FORM_SECTIONS.FINDINGS_TEXT,
            ),
            (
                SECONDARY_AUDITORS_ENTRY_FIXTURES,
                SECONDARY_AUDITORS_TEMPLATE,
                FORM_SECTIONS.SECONDARY_AUDITORS,
            ),
        ]

        for test_case in test_cases:
            with self.subTest():
                fixtures, template, section = test_case

                test_data = json.loads(fixtures.read_text(encoding="utf-8"))

                # add valid data to the workbook
                workbook = load_workbook(template, data_only=True)
                _set_by_name(workbook, "section_name", section)
                _add_entry(workbook, 0, test_data[0])
                # add invalid UEI to the workbook
                workbook = load_workbook(template, data_only=True)
                _set_by_name(workbook, "auditee_uei", ExcelFileHandlerViewTests.BAD_UEI)
                with NamedTemporaryFile(suffix=".xlsx") as tmp:
                    workbook.save(tmp.name)
                    tmp.seek(0)

                    with open(tmp.name, "rb") as excel_file:
                        response = self.client.post(
                            reverse(
                                f"audit:{section}",
                                kwargs={
                                    "report_id": sac.report_id,
                                    "form_section": section,
                                },
                            ),
                            data={"FILES": excel_file},
                        )

                        self.assertEqual(response.status_code, 400)
                        self.assertIn(
                            "The auditee UEI is not valid", response.json()["errors"][0]
                        )
                        self.assertFalse(ExcelFile.objects.exists())


class SingleAuditReportFileHandlerViewTests(TestCase):
    def test_login_required(self):
        """When an unauthenticated request is made"""

        response = self.client.post(
            reverse(
                "audit:SingleAuditReport",
                kwargs={"report_id": "12345"},
            )
        )

        self.assertTemplateUsed(response, "home.html")
        self.assertTrue(response.context["session_expired"])

    def test_bad_report_id_returns_403(self):
        """When a request is made for a malformed or nonexistent report_id, a 403 error should be returned"""
        user = baker.make(User)

        self.client.force_login(user)

        response = self.client.post(
            reverse(
                "audit:SingleAuditReport",
                kwargs={
                    "report_id": "this is not a report id",
                },
            )
        )

        self.assertEqual(response.status_code, 403)

    def test_inaccessible_audit_returns_403(self):
        """When a request is made for an audit that is inaccessible for this user, a 403 error should be returned"""
        user, sac = _make_user_and_sac()

        self.client.force_login(user)
        response = self.client.post(
            reverse(
                "audit:SingleAuditReport",
                kwargs={"report_id": sac.report_id},
            )
        )

        self.assertEqual(response.status_code, 403)

    def test_no_file_attached_returns_400(self):
        """When a request is made with no file attached, a 400 error should be returned"""
        user, sac = _make_user_and_sac()
        baker.make(Access, user=user, sac=sac)

        self.client.force_login(user)

        response = self.client.post(
            reverse(
                "audit:SingleAuditReport",
                kwargs={"report_id": sac.report_id},
            )
        )

        self.assertEqual(response.status_code, 400)

    @patch("audit.validators._scan_file")
    def test_valid_file_upload(self, mock_scan_file):
        """Test that uploading a valid SAR update the SAC accordingly"""
        sac = _mock_login_and_scan(
            self.client,
            mock_scan_file,
            report_id=_mock_gen_report_id(),
        )

        with open("audit/fixtures/basic.pdf", "rb") as pdf_file:
            response = self.client.post(
                reverse(
                    "audit:SingleAuditReport",
                    kwargs={
                        "report_id": sac.report_id,
                    },
                ),
                data={"FILES": pdf_file},
            )

            self.assertEqual(response.status_code, 302)

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be AUDIT_REPORT_PDF_UPDATED
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.AUDIT_REPORT_PDF_UPDATED,
        )

    @patch("audit.validators._scan_file")
    def test_valid_file_upload_for_additional_ueis(self, mock_scan_file):
        """When a valid Excel file is uploaded, the file should be stored and the SingleAuditChecklist should be updated to include the uploaded Additional UEIs data"""

        sac = _mock_login_and_scan(
            self.client,
            mock_scan_file,
            report_id=_mock_gen_report_id(),
        )
        test_data = json.loads(
            ADDITIONAL_UEIS_ENTRY_FIXTURES.read_text(encoding="utf-8")
        )

        # add valid data to the workbook
        workbook = load_workbook(ADDITIONAL_UEIS_TEMPLATE, data_only=True)
        _set_by_name(workbook, "auditee_uei", ExcelFileHandlerViewTests.GOOD_UEI)
        _set_by_name(workbook, "section_name", FORM_SECTIONS.ADDITIONAL_UEIS)
        _add_entry(workbook, 0, test_data[0])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)

            with open(tmp.name, "rb") as excel_file:
                response = self.client.post(
                    reverse(
                        f"audit:{FORM_SECTIONS.ADDITIONAL_UEIS}",
                        kwargs={
                            "report_id": sac.report_id,
                            "form_section": FORM_SECTIONS.ADDITIONAL_UEIS,
                        },
                    ),
                    data={"FILES": excel_file},
                )

                self.assertEqual(response.status_code, 302)

                updated_sac = SingleAuditChecklist.objects.get(pk=sac.id)

                self.assertEqual(
                    updated_sac.additional_ueis["AdditionalUEIs"]["auditee_uei"],
                    ExcelFileHandlerViewTests.GOOD_UEI,
                )

                self.assertEqual(
                    len(
                        updated_sac.additional_ueis["AdditionalUEIs"][
                            "additional_ueis_entries"
                        ]
                    ),
                    1,
                )

                additional_ueis_entries = updated_sac.additional_ueis["AdditionalUEIs"][
                    "additional_ueis_entries"
                ][0]

                self.assertEqual(
                    additional_ueis_entries["additional_uei"],
                    test_data[0]["additional_uei"],
                )

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be ADDITIONAL_UEIS_UPDATED
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.ADDITIONAL_UEIS_UPDATED,
        )

    @patch("audit.validators._scan_file")
    def test_valid_file_upload_for_additional_eins(self, mock_scan_file):
        """When a valid Excel file is uploaded, the file should be stored and the SingleAuditChecklist should be updated to include the uploaded Additional EINs data"""

        sac = _mock_login_and_scan(
            self.client,
            mock_scan_file,
            report_id=_mock_gen_report_id(),
        )
        test_data = json.loads(
            ADDITIONAL_EINS_ENTRY_FIXTURES.read_text(encoding="utf-8")
        )

        # add valid data to the workbook
        workbook = load_workbook(ADDITIONAL_EINS_TEMPLATE, data_only=True)
        _set_by_name(workbook, "auditee_uei", ExcelFileHandlerViewTests.GOOD_UEI)
        _set_by_name(workbook, "section_name", FORM_SECTIONS.ADDITIONAL_EINS)
        _add_entry(workbook, 0, test_data[0])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)

            with open(tmp.name, "rb") as excel_file:
                response = self.client.post(
                    reverse(
                        f"audit:{FORM_SECTIONS.ADDITIONAL_EINS}",
                        kwargs={
                            "report_id": sac.report_id,
                            "form_section": FORM_SECTIONS.ADDITIONAL_EINS,
                        },
                    ),
                    data={"FILES": excel_file},
                )

                self.assertEqual(response.status_code, 302)

                updated_sac = SingleAuditChecklist.objects.get(pk=sac.id)

                self.assertEqual(
                    updated_sac.additional_eins["AdditionalEINs"]["auditee_uei"],
                    ExcelFileHandlerViewTests.GOOD_UEI,
                )

                self.assertEqual(
                    len(
                        updated_sac.additional_eins["AdditionalEINs"][
                            "additional_eins_entries"
                        ]
                    ),
                    1,
                )

                additional_eins_entries = updated_sac.additional_eins["AdditionalEINs"][
                    "additional_eins_entries"
                ][0]

                self.assertEqual(
                    additional_eins_entries["additional_ein"],
                    test_data[0]["additional_ein"],
                )

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be ADDITIONAL_EINS_UPDATED
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.ADDITIONAL_EINS_UPDATED,
        )

    @patch("audit.validators._scan_file")
    def test_valid_file_upload_for_notes_to_sefa(self, mock_scan_file):
        """When a valid Excel file is uploaded, the file should be stored and the SingleAuditChecklist should be updated to include the uploaded Notes to SEFA data"""

        sac = _mock_login_and_scan(
            self.client,
            mock_scan_file,
            report_id=_mock_gen_report_id(),
        )
        test_data = json.loads(NOTES_TO_SEFA_ENTRY_FIXTURES.read_text(encoding="utf-8"))

        # add valid data to the workbook
        workbook = load_workbook(NOTES_TO_SEFA_TEMPLATE, data_only=True)
        _set_by_name(workbook, "auditee_uei", ExcelFileHandlerViewTests.GOOD_UEI)
        _set_by_name(workbook, "accounting_policies", "Mandatory notes")
        _set_by_name(workbook, "is_minimis_rate_used", "N")
        _set_by_name(workbook, "rate_explained", "More explanation.")
        _set_by_name(workbook, "section_name", FORM_SECTIONS.NOTES_TO_SEFA)
        _add_entry(workbook, 0, test_data[0])

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)

            with open(tmp.name, "rb") as excel_file:
                response = self.client.post(
                    reverse(
                        f"audit:{FORM_SECTIONS.NOTES_TO_SEFA}",
                        kwargs={
                            "report_id": sac.report_id,
                            "form_section": FORM_SECTIONS.NOTES_TO_SEFA,
                        },
                    ),
                    data={"FILES": excel_file},
                )

                self.assertEqual(response.status_code, 302)

                updated_sac = SingleAuditChecklist.objects.get(pk=sac.id)

                self.assertEqual(
                    updated_sac.notes_to_sefa["NotesToSefa"]["auditee_uei"],
                    ExcelFileHandlerViewTests.GOOD_UEI,
                )

                self.assertEqual(
                    len(
                        updated_sac.notes_to_sefa["NotesToSefa"][
                            "notes_to_sefa_entries"
                        ]
                    ),
                    1,
                )

                notes_to_sefa_entries = updated_sac.notes_to_sefa["NotesToSefa"][
                    "notes_to_sefa_entries"
                ][0]

                self.assertEqual(
                    notes_to_sefa_entries["note_title"],
                    test_data[0]["note_title"],
                )

        submission_events = SubmissionEvent.objects.filter(sac=sac)

        # the most recent event should be NOTES_TO_SEFA_UPDATED
        event_count = len(submission_events)
        self.assertGreaterEqual(event_count, 1)
        self.assertEqual(
            submission_events[event_count - 1].event,
            SubmissionEvent.EventType.NOTES_TO_SEFA_UPDATED,
        )


class EditSubmissionTest(TestCase):
    def setUp(self):
        """Setup factory, client, user, SAC, and URL"""
        self.factory = RequestFactory()
        self.client = Client()
        self.user = baker.make(User)
        self.sac = baker.make(
            SingleAuditChecklist, submission_status=STATUS.READY_FOR_CERTIFICATION
        )
        self.url = reverse(
            "audit:EditSubmission", kwargs={"report_id": self.sac.report_id}
        )
        self.client.force_login(self.user)
        baker.make(
            "audit.Access",
            sac=self.sac,
            user=self.user,
            role="certifying_auditee_contact",
        )
        self.session = self.client.session

    def test_redirects_to_singleauditchecklist(self):
        """Test that accessing edit submission redirects to SAC view"""
        response = self.client.get(self.url)
        self.assertRedirects(
            response, reverse("singleauditchecklist", args=[self.sac.report_id])
        )


class AuditorCertificationStep1ViewTests(TestCase):
    def setUp(self):
        """Setup client, user, SAC, and URL"""
        self.client = Client()
        self.user = baker.make(User)
        self.sac = baker.make(
            SingleAuditChecklist, submission_status=STATUS.READY_FOR_CERTIFICATION
        )
        self.url = reverse(
            "audit:AuditorCertification", kwargs={"report_id": self.sac.report_id}
        )
        self.client.force_login(self.user)
        baker.make(
            "audit.Access",
            sac=self.sac,
            user=self.user,
            role="certifying_auditor_contact",
        )
        self.session = self.client.session

    def test_get_redirects_if_status_not_ready_for_certification(self):
        """Test that GET redirects if SAC status is not READY_FOR_CERTIFICATION"""
        self.sac.submission_status = STATUS.IN_PROGRESS
        self.sac.save()
        response = self.client.get(self.url)
        self.assertRedirects(
            response, f"/audit/submission-progress/{self.sac.report_id}"
        )

    def test_get_renders_template_if_valid_state(self):
        """
        Test that GET renders the auditor certification setp 1 template when SAC
        is in a valid state.
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "audit/auditor-certification-step-1.html")
        self.assertIn("form", response.context)
        self.assertIsInstance(response.context["form"], AuditorCertificationStep1Form)

    def test_get_permission_denied_if_sac_not_found(self):
        """Test that POST redirects if SAC report_id is not found"""
        response = self.client.get(
            reverse("audit:AuditorCertification", kwargs={"report_id": "INVALID"})
        )
        self.assertEqual(response.status_code, 403)

    def test_post_redirects_if_status_not_ready_for_certification(self):
        """Test that POST redirects if SAC status is not READY_FOR_CERTIFICATION"""
        self.sac.submission_status = STATUS.IN_PROGRESS
        self.sac.save()
        response = self.client.post(self.url, {"field": "value"})
        self.assertRedirects(
            response, f"/audit/submission-progress/{self.sac.report_id}"
        )

    def test_post_valid_form(self):
        """
        Test that submitting a valid form updates the session
        and redirects correctly
        """
        self.session["AuditorCertificationStep1Session"] = {"field": "value"}
        self.session.save()

        form_data = {
            "is_OMB_limited": True,
            "is_auditee_responsible": True,
            "has_used_auditors_report": True,
            "has_no_auditee_procedures": True,
            "is_FAC_releasable": True,
        }

        response = self.client.post(self.url, form_data)

        self.assertEqual(response.status_code, 302)
        self.sac.refresh_from_db()

        self.assertRedirects(
            response,
            reverse("audit:AuditorCertificationConfirm", args=[self.sac.report_id]),
        )
        self.assertIn("AuditorCertificationStep1Session", self.client.session)
        self.assertEqual(
            self.client.session["AuditorCertificationStep1Session"], form_data
        )

    def test_post_invalid_form(self):
        """Test that submitting and invalid form renders an error template"""
        response = self.client.post(self.url, {"invalid_field": ""})
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "audit/auditor-certification-step-1.html")
        self.assertIn("form", response.context)
        self.assertIsInstance(response.context["form"], AuditorCertificationStep1Form)
        self.assertTrue(response.context["form"].errors)

    def test_post_permission_denied_if_sac_not_found(self):
        """Test that results in 403 if SAC report_id is not found"""
        response = self.client.post(
            reverse("audit:AuditorCertification", kwargs={"report_id": "INVALID"})
        )
        self.assertEqual(response.status_code, 403)


class AuditeeCertificationStep2ViewTests(TestCase):
    def setUp(self):
        """Setup client, user, SAC, and URL"""
        self.client = Client()
        self.user = baker.make(User)
        self.sac = baker.make(
            SingleAuditChecklist, submission_status=STATUS.AUDITOR_CERTIFIED
        )
        self.url = reverse(
            "audit:AuditeeCertificationConfirm",
            kwargs={"report_id": self.sac.report_id},
        )
        self.client.force_login(self.user)
        baker.make(
            "audit.Access",
            sac=self.sac,
            user=self.user,
            role="certifying_auditee_contact",
        )
        self.session = self.client.session

    def test_get_redirects_if_no_step_1(self):
        """Test that GET redirects if AuditeeCertificationStep1Session is missing"""
        response = self.client.get(self.url)
        self.assertRedirects(
            response, reverse("audit:AuditeeCertification", args=[self.sac.report_id])
        )

    def test_get_renders_template_if_valid_session(self):
        """
        Test that GET renders the Auditee certification step 2
        if session is valid
        """
        self.session["AuditeeCertificationStep1Session"] = {"field": "value"}
        self.session.save()
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "audit/auditee-certification-step-2.html")
        self.assertIn("form", response.context)
        self.assertIsInstance(response.context["form"], AuditeeCertificationStep2Form)

    def test_get_redirects_if_not_auditor_certified(self):
        """Test that GET will redirect if SAC status is not AUDITOR_CERTIFIED"""
        self.sac.submission_status = STATUS.IN_PROGRESS
        self.sac.save()
        self.session["AuditeeCertificationStep1Session"] = {"field": "value"}
        self.session.save()
        response = self.client.get(self.url)
        self.assertRedirects(
            response, f"/audit/submission-progress/{self.sac.report_id}"
        )

    @patch("audit.views.views.validate_auditee_certification_json")
    @patch("audit.views.views.sac_transition")
    def test_post_valid_form(self, mock_transition, mock_validate):
        """
        Test that submitting a valid Auditee Certification Form
        updates the SAC and redirects correctly
        """
        mock_transition.return_value = True
        mock_validate.return_value = {"auditee_certification": "validated_data"}

        self.session["AuditeeCertificationStep1Session"] = {"field": "value"}
        self.session.save()

        self.sac.submission_status = STATUS.AUDITOR_CERTIFIED
        self.sac.save()

        response = self.client.post(
            self.url,
            {
                "auditee_certification_date_signed": "2024-01-01",
                "auditee_name": "Test Auditee",
                "auditee_title": "Auditor",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.sac.refresh_from_db()
        mock_transition.assert_called_once_with(
            response.wsgi_request, self.sac, transition_to=STATUS.AUDITEE_CERTIFIED
        )
        self.assertRedirects(
            response, reverse("audit:SubmissionProgress", args=[self.sac.report_id])
        )

    def test_post_invalid_form(self):
        """
        Test that submitting an invalid Auditee Ceritifcation Form
        renders an error template
        """
        self.session["AuditeeCertificationStep1Session"] = None
        self.session.save()
        response = self.client.post(self.url, {"auditee_certification_date_signed": ""})

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "audit/auditee-certification-step-2.html")
        self.assertIn("form", response.context)
        self.assertIsInstance(response.context["form"], AuditeeCertificationStep2Form)
        self.assertTrue(response.context["form"].errors)

    def test_post_redirects_if_status_not_auditor_certified(self):
        """Test that POST will redirect if SAC submission status is not AUDITOR_CERTIFIED"""
        self.sac.submission_status = STATUS.IN_PROGRESS
        self.sac.save()
        self.session["AuditeeCertificationStep1Session"] = {"field": "value"}
        self.session.save()
        response = self.client.post(
            self.url, {"auditee_certification_date_signed": "2024-01-01"}
        )
        self.assertRedirects(
            response, f"/audit/submission-progress/{self.sac.report_id}"
        )

    def test_post_permission_denied_if_sac_not_found(self):
        """Test that POST will result in permission error is SAC report_id not found"""
        response = self.client.post(
            reverse(
                "audit:AuditeeCertificationConfirm", kwargs={"report_id": "INVALID"}
            )
        )
        self.assertEqual(response.status_code, 403)

    def test_single_audit_checklist_does_not_exist_exception(self):
        """Test that SAC does not exist renders a unique exception."""
        factory = RequestFactory()
        request = factory.get(reverse("audit:AuditeeCertification", args=["12345"]))

        with patch("audit.models.SingleAuditChecklist.objects.get") as mock_get:
            mock_get.side_effect = SingleAuditChecklist.DoesNotExist

            view = AuditeeCertificationStep2View.as_view()

            with self.assertRaises(PermissionDenied) as context:
                view(request, report_id="12345")

            self.assertEqual(
                str(context.exception), "You do not have access to this audit."
            )


class CrossValidationViewTests(TestCase):
    def setUp(self):
        """Setup client, user, SAC, and URL"""
        self.client = Client()
        self.user = baker.make(User)
        self.sac = baker.make(
            SingleAuditChecklist,
            report_id="test-report-id",
            submission_status=STATUS.IN_PROGRESS,
            general_information={"auditee_fiscal_period_end": "2024-12-31"},
        )
        self.url = reverse(
            "audit:CrossValidation", kwargs={"report_id": self.sac.report_id}
        )
        self.client.force_login(self.user)
        baker.make(
            "audit.Access",
            sac=self.sac,
            user=self.user,
            role="certifying_auditee_contact",
        )

    def test_get_view_renders_correct_template(self):
        """Test that GET renders cross-validation template with correct context"""
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "audit/cross-validation/cross-validation.html"
        )
        self.assertEqual(response.context["report_id"], self.sac.report_id)
        self.assertEqual(
            response.context["submission_status"], self.sac.submission_status
        )

    def test_get_view_permission_denied(self):
        """Test that GET returns 403 if SAC report_id does not exist"""
        url = reverse("audit:CrossValidation", args=["non-existent-id"])
        response = self.client.get(url)

        self.assertEqual(response.status_code, 403)  # PermissionDenied results in 403

    @patch("audit.models.SingleAuditChecklist.validate_full")
    def test_post_view_renders_results_template(self, mock_validate_full):
        """Test that POST with validation errors renders template with errors"""
        mock_validate_full.return_value = ["Error 1", "Error 2"]

        response = self.client.post(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(
            response, "audit/cross-validation/cross-validation-results.html"
        )
        self.assertEqual(response.context["report_id"], self.sac.report_id)
        self.assertEqual(response.context["errors"], ["Error 1", "Error 2"])
        mock_validate_full.assert_called_once()

    def test_post_view_permission_denied(self):
        """Test that POST returns 403 if SAC report_id does not exist"""
        url = reverse("audit:CrossValidation", args=["non-existent-id"])
        response = self.client.post(url)

        self.assertEqual(response.status_code, 403)


class RemoveSubmissionViewTests(TestCase):
    def setUp(self):
        """Setup client, user, valid/invalid statuses, and a report for each status"""
        self.client = Client()
        self.user = baker.make(User)
        self.client.force_login(self.user)

        # Static info
        self.template = "audit/remove-submission-in-progress.html"
        self.general_information = {
            "auditee_uei": "auditee_uei",
            "auditee_name": "auditee_name",
            "fiscal_year_end_date": "01/01/2022",
            "auditee_fiscal_period_end": "01/01/2022",
        }

        # Valid/invalid statuses. Create one report per status. Create one access per report.
        self.valid_removal_statuses = [
            STATUS.IN_PROGRESS,
            STATUS.READY_FOR_CERTIFICATION,
            STATUS.AUDITOR_CERTIFIED,
            STATUS.AUDITEE_CERTIFIED,
            STATUS.CERTIFIED,
        ]
        self.invalid_removal_statuses = [
            STATUS.SUBMITTED,
            STATUS.DISSEMINATED,
            STATUS.FLAGGED_FOR_REMOVAL,
        ]
        self.reports = [
            baker.make(
                SingleAuditChecklist,
                report_id=f"test-report-id--{status}",
                submission_status=status,
                general_information=self.general_information,
            )
            for status in self.valid_removal_statuses + self.invalid_removal_statuses
        ]
        for report in self.reports:
            baker.make(
                "audit.Access",
                email=self.user.email,
                sac=report,
                user=self.user,
                role="certifying_auditor_contact",
            )

    def test_get_view_permission_denied(self):
        url = reverse(
            "audit:RemoveSubmissionInProgress", kwargs={"report_id": "non-existent-id"}
        )
        response = self.client.get(url)

        self.assertEqual(response.status_code, 403)

    def test_post_view_permission_denied(self):
        """Test that POST returns 403 if SAC report_id does not exist"""
        url = reverse(
            "audit:RemoveSubmissionInProgress", kwargs={"report_id": "non-existent-id"}
        )
        response = self.client.post(url)

        self.assertEqual(response.status_code, 403)

    def test_get_view_renders_correct_template(self):
        """
        Test that GET correctly filters out reports of invalid status.
        """
        # For all possible report statuses, get the associated report and report_id. Make a GET request to the removal URL.
        for status in self.valid_removal_statuses + self.invalid_removal_statuses:
            report = next(
                (x for x in self.reports if x.submission_status == status), None
            )
            report_id = report.report_id
            url_removal = reverse(
                "audit:RemoveSubmissionInProgress", kwargs={"report_id": report_id}
            )
            response = self.client.get(url_removal, follow=True)
            url_after = response.request.get("PATH_INFO")

            # If the GET request was made with valid report status, expect the correct template to load.
            # Otherwise, expect to be redirected to the submission checklist.
            if status in self.valid_removal_statuses:
                self.assertEqual(response.status_code, 200)
                self.assertTemplateUsed(response, self.template)
            else:
                url_checklist = reverse(
                    "audit:SubmissionProgress", kwargs={"report_id": report_id}
                )
                self.assertEqual(url_after, url_checklist)

    def test_post_view_flags_correct_reports(self):
        """
        Test that POST correctly filters out reports of invalid status, and flags reports of valid status for removal.
        """
        url_audit_submisisons = reverse("audit:MySubmissions")

        # For all possible report statuses, get the associated report and report_id. Make a POST request to the removal URL.
        for status in self.valid_removal_statuses + self.invalid_removal_statuses:
            report = next(
                (x for x in self.reports if x.submission_status == status), None
            )
            report_id = report.report_id
            url_removal = reverse(
                "audit:RemoveSubmissionInProgress", kwargs={"report_id": report_id}
            )
            response = self.client.post(url_removal, follow=True)
            report_after = SingleAuditChecklist.objects.get(report_id=report_id)
            url_after = response.request.get("PATH_INFO")

            # If the POST request was made with valid report status, expect the audit submisisons page to load. Verify that the report has been flagged.
            # Otherwise, expect to be redirected to the submission checklist. Verify the report was not flagged.
            if status in self.valid_removal_statuses:
                self.assertEqual(url_after, url_audit_submisisons)
                self.assertEqual(
                    report_after.submission_status, STATUS.FLAGGED_FOR_REMOVAL
                )
            else:
                url_checklist = reverse(
                    "audit:SubmissionProgress", kwargs={"report_id": report_id}
                )
                self.assertEqual(url_after, url_checklist)
                self.assertEqual(report_after.submission_status, status)
