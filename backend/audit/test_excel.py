import json

from django.core.exceptions import ValidationError
from django.test import SimpleTestCase
from openpyxl import load_workbook
from openpyxl.cell import Cell

from audit.excel import (
    ExcelExtractionError,
    extract_federal_awards,
    extract_findings_text,
    extract_findings_uniform_guidance,
    extract_corrective_action_plan,
    extract_additional_ueis,
    extract_secondary_auditors,
    extract_notes_to_sefa,
    federal_awards_field_mapping,
    findings_text_field_mapping,
    findings_uniform_guidance_field_mapping,
    corrective_action_field_mapping,
    additional_ueis_field_mapping,
    secondary_auditors_field_mapping,
    notes_to_sefa_field_mapping,
    federal_awards_column_mapping,
    findings_text_column_mapping,
    findings_uniform_guidance_column_mapping,
    corrective_action_column_mapping,
    additional_ueis_column_mapping,
    secondary_auditors_column_mapping,
    notes_to_sefa_column_mapping,
)
from audit.validators import (
    validate_additional_ueis_json,
    validate_federal_award_json,
    validate_corrective_action_plan_json,
    validate_findings_text_json,
    validate_findings_uniform_guidance_json,
    validate_secondary_auditors_json,
    validate_notes_to_sefa_json,
)
from audit.fixtures.excel import (
    FEDERAL_AWARDS_TEMPLATE,
    FEDERAL_AWARDS_ENTRY_FIXTURES,
    CORRECTIVE_ACTION_PLAN_TEMPLATE,
    CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES,
    SECONDARY_AUDITORS_ENTRY_FIXTURES,
    FINDINGS_TEXT_TEMPLATE,
    FINDINGS_TEXT_ENTRY_FIXTURES,
    FINDINGS_UNIFORM_GUIDANCE_TEMPLATE,
    FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES,
    ADDITIONAL_UEIS_TEMPLATE,
    ADDITIONAL_UEIS_ENTRY_FIXTURES,
    SECONDARY_AUDITORS_TEMPLATE,
    NOTES_TO_SEFA_TEMPLATE,
    NOTES_TO_SEFA_ENTRY_FIXTURES,
)

# Simplest way to create a new copy of simple case rather than getting
# references to things used by other tests:
jsoncopy = lambda v: json.loads(json.dumps(v))


def _set_by_name(workbook, name, value, row_offset=0):
    definition = workbook.defined_names[name]

    sheet_title, cell_coord = next(definition.destinations)

    sheet = workbook[sheet_title]
    cell_range = sheet[cell_coord]

    if isinstance(cell_range, Cell):
        cell_range.value = value
    else:
        cell_range[row_offset][0].value = value


def _add_entry(workbook, row_offset, entry):
    for key, value in entry.items():
        _set_by_name(workbook, key, value, row_offset)


class FederalAwardsExcelTests(SimpleTestCase):
    GOOD_UEI = "AAA123456BBB"
    TEST_DATA = json.loads(FEDERAL_AWARDS_ENTRY_FIXTURES.read_text(encoding="utf-8"))

    def test_template_has_named_ranges(self):
        """Test that the FederalAwardsExpended Excel template contains the expected named ranges"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        for name in federal_awards_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in federal_awards_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_single_federal_awards_entry(self):
        """Test that extraction and validation succeed when there is a single federal awards entry"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", FederalAwardsExcelTests.GOOD_UEI)
        _set_by_name(workbook, "amount_expended", 100)
        _add_entry(workbook, 0, FederalAwardsExcelTests.TEST_DATA[0])

        federal_awards = extract_federal_awards(workbook)

        validate_federal_award_json(federal_awards)

    def test_multiple_federal_awards_entries(self):
        """Test that extraction and validation succeed when there are multiple federal awards entries"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", FederalAwardsExcelTests.GOOD_UEI)
        _set_by_name(workbook, "amount_expended", 200)
        for index, entry in enumerate(FederalAwardsExcelTests.TEST_DATA):
            _add_entry(workbook, index, entry)

        federal_awards = extract_federal_awards(workbook)

        validate_federal_award_json(federal_awards)

    def test_partial_federal_awards_entry(self):
        """Test that extraction succeeds and validation fails when there are partial federal awards entries"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", FederalAwardsExcelTests.GOOD_UEI)
        _set_by_name(workbook, "total_amount_expended", 200)

        entry = jsoncopy(FederalAwardsExcelTests.TEST_DATA[0])
        del entry["cluster_name"]

        _add_entry(workbook, 0, entry)

        federal_awards = extract_federal_awards(workbook)

        self.assertRaises(ValidationError, validate_federal_award_json, federal_awards)

    def test_federal_awards_type_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_uei", FederalAwardsExcelTests.GOOD_UEI)
        _set_by_name(workbook, "total_amount_expended", 200)
        _add_entry(workbook, 0, FederalAwardsExcelTests.TEST_DATA[0])

        test_cases = [
            ("auditee_uei", 123456789123),
            ("total_amount_expended", "not a number"),
            ("amount_expended", "not a  number"),
            ("cluster_name", 123),
            ("is_direct", 123),
            ("is_passed", 123),
            ("subrecipient_amount", "not a number"),
            ("program_name", 123),
            ("loan_balance_at_audit_period_end", "not a number"),
            ("is_guaranteed", 123),
            ("is_major", 123),
            ("audit_report_type", 123),
            ("number_of_audit_findings", "not a number"),
            ("federal_agency_prefix", 10),
            ("three_digit_extension", "001"),
            ("state_cluster_name", 123),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                federal_awards = extract_federal_awards(workbook)

                self.assertRaises(
                    ValidationError, validate_federal_award_json, federal_awards
                )

    def test_federal_awards_custom_formatters(self):
        """Test that custom federal awards field formatters raise the expected error type when data is malformed"""
        workbook = load_workbook(FEDERAL_AWARDS_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_uei", FederalAwardsExcelTests.GOOD_UEI)
        _set_by_name(workbook, "amount_expended", 200)
        _add_entry(workbook, 0, FederalAwardsExcelTests.TEST_DATA[0])

        test_cases = [
            ("passthrough_name", 0),
            ("passthrough_identifying_number", 0),
        ]

        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                self.assertRaises(
                    ExcelExtractionError, extract_federal_awards, workbook
                )


class CorrectiveActionPlanExcelTests(SimpleTestCase):
    GOOD_UEI = "AAA123456BBB"
    TOO_SHORT_UEI = "AAA123456"
    TEST_DATA = json.loads(
        CORRECTIVE_ACTION_PLAN_ENTRY_FIXTURES.read_text(encoding="utf-8")
    )

    def test_template_has_named_ranges(self):
        """Test that the CorrectiveActionPlan Excel template contains the expected named ranges"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        for name in corrective_action_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in corrective_action_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_single_corrective_action_plan_entry(self):
        """Test that extraction and validation succeed when there is a single corrective action plan entry"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", CorrectiveActionPlanExcelTests.GOOD_UEI)

        _add_entry(workbook, 0, CorrectiveActionPlanExcelTests.TEST_DATA[0])

        corrective_action_plan = extract_corrective_action_plan(workbook)

        validate_corrective_action_plan_json(corrective_action_plan)

    def test_multiple_corrective_action_plan_entries(self):
        """Test that extraction and validation succeed when there are multiple corrective action plan entries"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", CorrectiveActionPlanExcelTests.GOOD_UEI)

        for index, entry in enumerate(CorrectiveActionPlanExcelTests.TEST_DATA):
            _add_entry(workbook, index, entry)

        corrective_action_plan = extract_corrective_action_plan(workbook)

        validate_corrective_action_plan_json(corrective_action_plan)

    def test_partial_corrective_action_plan_entry(self):
        """Test that extraction succeeds and validation fails when there are partial corrective action plan entries"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", CorrectiveActionPlanExcelTests.GOOD_UEI)

        entry = jsoncopy(CorrectiveActionPlanExcelTests.TEST_DATA[0])
        del entry["planned_action"]

        _add_entry(workbook, 0, entry)

        corrective_action_plan = extract_corrective_action_plan(workbook)

        self.assertRaises(
            ValidationError,
            validate_corrective_action_plan_json,
            corrective_action_plan,
        )

    def test_corrective_action_plan_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(CORRECTIVE_ACTION_PLAN_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_uei", CorrectiveActionPlanExcelTests.GOOD_UEI)
        _add_entry(workbook, 0, CorrectiveActionPlanExcelTests.TEST_DATA[0])

        test_cases = [
            ("auditee_uei", CorrectiveActionPlanExcelTests.TOO_SHORT_UEI),
            ("contains_chart_or_table", "not a boolean"),
            ("planned_action", 0),
            ("reference_number", 0),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                corrective_action_plan = extract_corrective_action_plan(workbook)

                self.assertRaises(
                    ValidationError,
                    validate_corrective_action_plan_json,
                    corrective_action_plan,
                )


class FindingsUniformGuidanceExcelTests(SimpleTestCase):
    GOOD_UEI = "AAA123456BBB"
    TEST_DATA = json.loads(
        FINDINGS_UNIFORM_GUIDANCE_ENTRY_FIXTURES.read_text(encoding="utf-8")
    )

    def test_template_has_named_ranges(self):
        """Test that the FindingsUniformGuidance Excel template contains the expected named ranges"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        for name in findings_uniform_guidance_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in findings_uniform_guidance_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_single_findings_uniform_guidance_entry(self):
        """Test that extraction and validation succeed when there is a single findings uniform guidance entry"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        _set_by_name(
            workbook, "auditee_uei", FindingsUniformGuidanceExcelTests.GOOD_UEI
        )
        _add_entry(workbook, 0, FindingsUniformGuidanceExcelTests.TEST_DATA[0])

        findings = extract_findings_uniform_guidance(workbook)

        validate_findings_uniform_guidance_json(findings)

    def test_multiple_findings_uniform_guidance_entries(self):
        """Test that extraction and validation succeed when there are multiple findings uniform guidance entries"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        _set_by_name(
            workbook, "auditee_uei", FindingsUniformGuidanceExcelTests.GOOD_UEI
        )
        for index, entry in enumerate(FindingsUniformGuidanceExcelTests.TEST_DATA):
            _add_entry(workbook, index, entry)

        findings = extract_findings_uniform_guidance(workbook)

        validate_findings_uniform_guidance_json(findings)

    def test_partial_findings_uniform_guidance_entry(self):
        """Test that extraction succeeds and validation fails when there are partial findings uniform guidance entries"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        _set_by_name(
            workbook, "auditee_uei", FindingsUniformGuidanceExcelTests.GOOD_UEI
        )

        entry = jsoncopy(FindingsUniformGuidanceExcelTests.TEST_DATA[0])
        del entry["reference_number"]

        _add_entry(workbook, 0, entry)

        findings = extract_findings_uniform_guidance(workbook)

        self.assertRaises(
            ValidationError, validate_findings_uniform_guidance_json, findings
        )

    def test_findings_uniform_guidance_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(FINDINGS_UNIFORM_GUIDANCE_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(
            workbook, "auditee_uei", FindingsUniformGuidanceExcelTests.GOOD_UEI
        )
        _add_entry(workbook, 0, FindingsUniformGuidanceExcelTests.TEST_DATA[0])

        test_cases = [
            ("auditee_uei", 123456789123),
            ("reference_number", 0),
            ("program_name", 123),
            ("federal_agency_prefix", 10),
            ("three_digit_extension", "001"),
            ("prior_references", 123),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                findings = extract_findings_uniform_guidance(workbook)

                self.assertRaises(
                    ValidationError, validate_findings_uniform_guidance_json, findings
                )


class FindingsTextExcelTests(SimpleTestCase):
    GOOD_UEI = "AAA123456BBB"
    TEST_DATA = json.loads(FINDINGS_TEXT_ENTRY_FIXTURES.read_text(encoding="utf-8"))

    def test_template_has_named_ranges(self):
        """Test that the FindingsText Excel template contains the expected named ranges"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        for name in findings_text_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in findings_text_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_single_findings_text_entry(self):
        """Test that extraction and validation succeed when there is a single findings text entry"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", FindingsTextExcelTests.GOOD_UEI)
        _add_entry(workbook, 0, FindingsTextExcelTests.TEST_DATA[0])

        findings = extract_findings_text(workbook)

        validate_findings_text_json(findings)

    def test_multiple_findings_text_entries(self):
        """Test that extraction and validation succeed when there are multiple findings text entries"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", FindingsTextExcelTests.GOOD_UEI)
        for index, entry in enumerate(FindingsTextExcelTests.TEST_DATA):
            _add_entry(workbook, index, entry)

        findings = extract_findings_text(workbook)

        validate_findings_text_json(findings)

    def test_partial_findings_text_entry(self):
        """Test that extraction succeeds and validation fails when there are partial findings text entries"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", FindingsTextExcelTests.GOOD_UEI)

        entry = jsoncopy(FindingsTextExcelTests.TEST_DATA[0])
        del entry["text_of_finding"]

        _add_entry(workbook, 0, entry)

        findings = extract_findings_text(workbook)

        self.assertRaises(ValidationError, validate_findings_text_json, findings)

    def test_findings_text_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(FINDINGS_TEXT_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_uei", FindingsTextExcelTests.GOOD_UEI)
        _add_entry(workbook, 0, FindingsTextExcelTests.TEST_DATA[0])

        test_cases = [
            ("auditee_uei", 123456789123),
            ("reference_number", 0),
            ("contains_chart_or_table", "not a boolean"),
            ("text_of_finding", 10.001),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                findings = extract_findings_text(workbook)

                self.assertRaises(
                    ValidationError, validate_findings_text_json, findings
                )


class AdditionalUeisExcelTests(SimpleTestCase):
    GOOD_UEI = "AAA123456BBB"
    TEST_DATA = json.loads(ADDITIONAL_UEIS_ENTRY_FIXTURES.read_text(encoding="utf-8"))

    def test_template_has_named_ranges(self):
        """Test that the AdditionalUEIs Excel template contains the expected named ranges"""
        workbook = load_workbook(ADDITIONAL_UEIS_TEMPLATE, data_only=True)

        for name in additional_ueis_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in additional_ueis_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_single_additional_ueis_entry(self):
        """Test that extraction and validation succeed when there is a single additional ueis entry"""
        workbook = load_workbook(ADDITIONAL_UEIS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", AdditionalUeisExcelTests.GOOD_UEI)
        _add_entry(workbook, 0, AdditionalUeisExcelTests.TEST_DATA[0])

        additional_ueis = extract_additional_ueis(workbook)

        validate_additional_ueis_json(additional_ueis)

    def test_multiple_additional_ueis_entries(self):
        """Test that extraction and validation succeed when there are multiple additional ueis entries"""
        workbook = load_workbook(ADDITIONAL_UEIS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", AdditionalUeisExcelTests.GOOD_UEI)
        for index, entry in enumerate(AdditionalUeisExcelTests.TEST_DATA):
            _add_entry(workbook, index, entry)

        additional_ueis = extract_additional_ueis(workbook)

        validate_additional_ueis_json(additional_ueis)

    def test_additional_ueis_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(ADDITIONAL_UEIS_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_uei", AdditionalUeisExcelTests.GOOD_UEI)
        _add_entry(workbook, 0, AdditionalUeisExcelTests.TEST_DATA[0])

        test_cases = [
            ("auditee_uei", 123456789123),
            ("additional_uei", 123456789123),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                additional_ueis = extract_additional_ueis(workbook)

                self.assertRaises(
                    ValidationError, validate_additional_ueis_json, additional_ueis
                )


class SecondaryAuditorsExcelTests(SimpleTestCase):
    GOOD_UEI = "AAA123456BBB"
    TEST_DATA = json.loads(
        SECONDARY_AUDITORS_ENTRY_FIXTURES.read_text(encoding="utf-8")
    )

    def test_template_has_named_ranges(self):
        """Test that the SecondaryAuditors Excel template contains the expected named ranges"""
        workbook = load_workbook(SECONDARY_AUDITORS_TEMPLATE, data_only=True)

        for name in secondary_auditors_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in secondary_auditors_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_single_secondary_auditors_entry(self):
        """Test that extraction and validation succeed when there is a single secondary auditors entry"""
        workbook = load_workbook(SECONDARY_AUDITORS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", SecondaryAuditorsExcelTests.GOOD_UEI)
        _add_entry(workbook, 0, SecondaryAuditorsExcelTests.TEST_DATA[0])

        secondary_auditors = extract_secondary_auditors(workbook)

        validate_secondary_auditors_json(secondary_auditors)

    def test_multiple_secondary_auditors_entries(self):
        """Test that extraction and validation succeed when there are multiple secondary auditors entries"""
        workbook = load_workbook(SECONDARY_AUDITORS_TEMPLATE, data_only=True)

        _set_by_name(workbook, "auditee_uei", SecondaryAuditorsExcelTests.GOOD_UEI)
        for index, entry in enumerate(SecondaryAuditorsExcelTests.TEST_DATA):
            _add_entry(workbook, index, entry)

        secondary_auditors = extract_secondary_auditors(workbook)

        validate_secondary_auditors_json(secondary_auditors)

    def test_secondary_auditors_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(SECONDARY_AUDITORS_TEMPLATE, data_only=True)

        # add valid data to the workbook
        _set_by_name(workbook, "auditee_uei", SecondaryAuditorsExcelTests.GOOD_UEI)
        _add_entry(workbook, 0, SecondaryAuditorsExcelTests.TEST_DATA[0])

        test_cases = [
            ("auditee_uei", 123456789123),
            ("secondary_auditor_name", False),
            ("secondary_auditor_ein", 12345678),
            ("secondary_auditor_address", True),
            ("secondary_auditor_city", 0),
            ("secondary_auditor_state", "Of Mind"),
            ("secondary_auditor_zip", "91919999"),
            ("secondary_auditor_contact_name", 11),
            ("secondary_auditor_contact_title", 4),
            ("secondary_auditor_contact_phone", "333-333-55555"),
            ("secondary_auditor_contact_email", "tory.audi$bae.com"),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                secondary_auditors = extract_secondary_auditors(workbook)

                self.assertRaises(
                    ValidationError,
                    validate_secondary_auditors_json,
                    secondary_auditors,
                )


class NotesToSefaExcelTests(SimpleTestCase):
    GOOD_UEI = "AAA123456BBB"
    TEST_DATA = json.loads(NOTES_TO_SEFA_ENTRY_FIXTURES.read_text(encoding="utf-8"))

    def test_template_has_named_ranges(self):
        """Test that the NotesToSefa Excel template contains the expected named ranges"""
        workbook = load_workbook(NOTES_TO_SEFA_TEMPLATE, data_only=True)

        for name in notes_to_sefa_field_mapping.keys():
            self.assertIsNotNone(workbook.defined_names[name])

        for name in notes_to_sefa_column_mapping:
            self.assertIsNotNone(workbook.defined_names[name])

    def test_single_notes_to_sefa_entry(self):
        """Test that extraction and validation succeed when there is a single notes to sefa entry"""
        workbook = load_workbook(NOTES_TO_SEFA_TEMPLATE, data_only=True)

        NotesToSefaExcelTests._set_required_fields(workbook)
        _add_entry(workbook, 0, NotesToSefaExcelTests.TEST_DATA[0])

        notes_to_sefa = extract_notes_to_sefa(workbook)

        validate_notes_to_sefa_json(notes_to_sefa)

    def test_multiple_notes_to_sefa_entries(self):
        """Test that extraction and validation succeed when there are multiple notes to sefa entries"""
        workbook = load_workbook(NOTES_TO_SEFA_TEMPLATE, data_only=True)

        NotesToSefaExcelTests._set_required_fields(workbook)
        for index, entry in enumerate(NotesToSefaExcelTests.TEST_DATA):
            _add_entry(workbook, index, entry)

        notes_to_sefa = extract_notes_to_sefa(workbook)

        validate_notes_to_sefa_json(notes_to_sefa)

    def test_notes_to_sefa_checking(self):
        """Test that extraction succeeds and validation fails when fields are of the wrong data type"""
        workbook = load_workbook(NOTES_TO_SEFA_TEMPLATE, data_only=True)

        # add valid data to the workbook
        NotesToSefaExcelTests._set_required_fields(workbook)
        _add_entry(workbook, 0, NotesToSefaExcelTests.TEST_DATA[0])

        test_cases = [
            ("auditee_uei", 123456789123),
        ]

        # validate that each test_case appropriately checks the data type
        for field_name, value in test_cases:
            with self.subTest():
                _set_by_name(workbook, field_name, value)

                notes_to_sefa = extract_notes_to_sefa(workbook)

                self.assertRaises(
                    ValidationError, validate_notes_to_sefa_json, notes_to_sefa
                )

    def _set_required_fields(workbook):
        _set_by_name(workbook, "auditee_uei", NotesToSefaExcelTests.GOOD_UEI)
        _set_by_name(workbook, "accounting_policies", "Mandatory notes")
        _set_by_name(workbook, "is_minimis_rate_used", "Y")
        _set_by_name(workbook, "rate_explained", "More explanation.")
