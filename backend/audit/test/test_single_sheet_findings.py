import openpyxl
import os

file_path = "/src/audit/test/2024-10-GSAFAC-0000386512--FindingsUniformGuidance (2).xlsx"
expected_sheet_prefix = "form"
required_columns = [
    "Award Reference",
    "Audit Finding Reference Number",
    "Type(s) of Compliance Requirement(s)",
    "Modified Opinion",
    "Other Matters",
    "Material Weakness",
    "Significant Deficiency",
    "Other Findings",
    "Questioned Costs",
    "Repeat Findings from Prior Year",
    "If Repeat Finding, provide Prior Year Audit Finding Reference Number(s)",
    "Is Findings Combination Valid? (Read Only)",
]

def validate_findings_sheet(path):
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return

    wb = openpyxl.load_workbook(path)
    matched_sheet = None
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower().startswith(expected_sheet_prefix):
            matched_sheet = sheet_name
            break

    if not matched_sheet:
        print("Could not find the Findings Uniform Guidance sheet.")
        print("Available sheets:", wb.sheetnames)
        return

    ws = wb[matched_sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    missing = [col for col in required_columns if col not in headers]
    if missing:
        print("Missing required columns:")
        for col in missing:
            print(f"- {col}")
    else:
        print("Findings sheet structure looks good.")

if __name__ == "__main__":
    validate_findings_sheet(file_path)