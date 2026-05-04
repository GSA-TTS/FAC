import openpyxl
import os

file_path = "/src/audit/test/2024-10-GSAFAC-0000386512--FindingsText (1) (1).xlsx"
required_columns = [
    "Audit Finding Reference Number",
    "Text of the Audit Finding",
]

def validate_sheet(path):
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return

    wb = openpyxl.load_workbook(path)
    matched_sheet = None
    for s in wb.sheetnames:
        if s.strip().lower() == "form":
            matched_sheet = s
            break

    if not matched_sheet:
        print("Could not find a sheet for Findings Text.")
        print("Available sheets:", wb.sheetnames)
        return

    ws = wb[matched_sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    missing = [col for col in required_columns if col not in headers]
    if missing:
        print("Missing required columns:")
        for col in missing:
            print(f"- {col}")
    else:
        print("Findings Text sheet structure looks good.")

if __name__ == "__main__":
    validate_sheet(file_path)
